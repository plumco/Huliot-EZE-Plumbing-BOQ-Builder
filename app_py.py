import streamlit as st
import pandas as pd
from io import BytesIO
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import uuid

st.set_page_config(page_title="Huliot EZE BOQ", page_icon="⚙", layout="wide")

st.markdown("""
<style>
#MainMenu,footer,header{visibility:hidden}
.block-container{padding-top:0.6rem;padding-bottom:1rem}
.stButton>button{font-weight:700;border-radius:8px;font-size:12px}
.stTextInput>div>div>input{font-size:13px}
.stSelectbox>div>div{font-size:13px}
.stTabs [data-baseweb="tab"]{font-weight:700;font-size:13px}
.stTabs [data-baseweb="tab-list"]{gap:6px;border-bottom:2px solid #E2E8F0}
</style>
""", unsafe_allow_html=True)

# ── PRODUCT DATA ──
def _p(code, desc, ptype, dn, price, line, sub=""):
    return {"code":code,"desc":desc,"type":ptype,"sub":sub,"dn":dn,"price":price,"line":line}

PRODUCTS = [
    # ════════════════════════════════════════
    # HT PRO — PIPES
    # ════════════════════════════════════════
    # S/S 250mm
    _p("5404000025-i","S/S Pipe L=250mm","Pipe",40,77,"HT Pro","S/S 250mm"),
    _p("5405000025-i","S/S Pipe L=250mm","Pipe",50,88,"HT Pro","S/S 250mm"),
    _p("5407500025-i","S/S Pipe L=250mm","Pipe",75,147,"HT Pro","S/S 250mm"),
    _p("5401100025-i","S/S Pipe L=250mm","Pipe",110,243,"HT Pro","S/S 250mm"),
    _p("5401200025-i","S/S Pipe L=250mm","Pipe",125,392,"HT Pro","S/S 250mm"),
    _p("5401600025-i","S/S Pipe L=250mm","Pipe",160,483,"HT Pro","S/S 250mm"),
    # S/S 500mm
    _p("54040050-i","S/S Pipe L=500mm","Pipe",40,113,"HT Pro","S/S 500mm"),
    _p("54050050-i","S/S Pipe L=500mm","Pipe",50,147,"HT Pro","S/S 500mm"),
    _p("54075050-i","S/S Pipe L=500mm","Pipe",75,252,"HT Pro","S/S 500mm"),
    _p("540110050-i","S/S Pipe L=500mm","Pipe",110,411,"HT Pro","S/S 500mm"),
    _p("540125050-i","S/S Pipe L=500mm","Pipe",125,664,"HT Pro","S/S 500mm"),
    _p("540160050-i","S/S Pipe L=500mm","Pipe",160,812,"HT Pro","S/S 500mm"),
    # S/S 1000mm
    _p("54040100-i","S/S Pipe L=1000mm","Pipe",40,204,"HT Pro","S/S 1m"),
    _p("54050100-i","S/S Pipe L=1000mm","Pipe",50,266,"HT Pro","S/S 1m"),
    _p("54075100-i","S/S Pipe L=1000mm","Pipe",75,462,"HT Pro","S/S 1m"),
    _p("540110100-i","S/S Pipe L=1000mm","Pipe",110,747,"HT Pro","S/S 1m"),
    _p("540125100-i","S/S Pipe L=1000mm","Pipe",125,1212,"HT Pro","S/S 1m"),
    _p("540160100-i","S/S Pipe L=1000mm","Pipe",160,1474,"HT Pro","S/S 1m"),
    # S/S 1500mm
    _p("54040150-i","S/S Pipe L=1500mm","Pipe",40,294,"HT Pro","S/S 1.5m"),
    _p("54050150-i","S/S Pipe L=1500mm","Pipe",50,384,"HT Pro","S/S 1.5m"),
    _p("54075150-i","S/S Pipe L=1500mm","Pipe",75,672,"HT Pro","S/S 1.5m"),
    _p("540110150-i","S/S Pipe L=1500mm","Pipe",110,1085,"HT Pro","S/S 1.5m"),
    _p("540125150-i","S/S Pipe L=1500mm","Pipe",125,1759,"HT Pro","S/S 1.5m"),
    _p("540160150-i","S/S Pipe L=1500mm","Pipe",160,2134,"HT Pro","S/S 1.5m"),
    # S/S 2000mm
    _p("54040200-i","S/S Pipe L=2000mm","Pipe",40,385,"HT Pro","S/S 2m"),
    _p("54050200-i","S/S Pipe L=2000mm","Pipe",50,503,"HT Pro","S/S 2m"),
    _p("54075200-i","S/S Pipe L=2000mm","Pipe",75,882,"HT Pro","S/S 2m"),
    _p("540110200-i","S/S Pipe L=2000mm","Pipe",110,1421,"HT Pro","S/S 2m"),
    _p("540125200-i","S/S Pipe L=2000mm","Pipe",125,2307,"HT Pro","S/S 2m"),
    _p("540160200-i","S/S Pipe L=2000mm","Pipe",160,2796,"HT Pro","S/S 2m"),
    # S/S 3000mm
    _p("54040300-i","S/S Pipe L=3000mm","Pipe",40,567,"HT Pro","S/S 3m"),
    _p("54050300-i","S/S Pipe L=3000mm","Pipe",50,739,"HT Pro","S/S 3m"),
    _p("54075300-i","S/S Pipe L=3000mm","Pipe",75,1304,"HT Pro","S/S 3m"),
    _p("540110300-i","S/S Pipe L=3000mm","Pipe",110,1751,"HT Pro","S/S 3m"),
    _p("540125300-i","S/S Pipe L=3000mm","Pipe",125,3400,"HT Pro","S/S 3m"),
    _p("540160300-i","S/S Pipe L=3000mm","Pipe",160,4116,"HT Pro","S/S 3m"),
    _p("540200300-i","S/S Pipe L=3000mm","Pipe",200,8303,"HT Pro","S/S 3m"),
    # D/S 500mm
    _p("5404040050-i","D/S Pipe L=500mm","Pipe",40,134,"HT Pro","D/S 500mm"),
    _p("540505050-i","D/S Pipe L=500mm","Pipe",50,175,"HT Pro","D/S 500mm"),
    _p("5407575050-i","D/S Pipe L=500mm","Pipe",75,292,"HT Pro","D/S 500mm"),
    _p("5401111050-i","D/S Pipe L=500mm","Pipe",110,484,"HT Pro","D/S 500mm"),
    _p("5401212050-i","D/S Pipe L=500mm","Pipe",125,782,"HT Pro","D/S 500mm"),
    _p("5401616050-i","D/S Pipe L=500mm","Pipe",160,964,"HT Pro","D/S 500mm"),
    # D/S 1000mm
    _p("5404040100-i","D/S Pipe L=1000mm","Pipe",40,224,"HT Pro","D/S 1m"),
    _p("5405050100-i","D/S Pipe L=1000mm","Pipe",50,293,"HT Pro","D/S 1m"),
    _p("5407575100-i","D/S Pipe L=1000mm","Pipe",75,502,"HT Pro","D/S 1m"),
    _p("5401111100-i","D/S Pipe L=1000mm","Pipe",110,821,"HT Pro","D/S 1m"),
    _p("5401212100-i","D/S Pipe L=1000mm","Pipe",125,1328,"HT Pro","D/S 1m"),
    _p("5401616100-i","D/S Pipe L=1000mm","Pipe",160,1624,"HT Pro","D/S 1m"),
    # D/S 1500mm
    _p("5404040150-i","D/S Pipe L=1500mm","Pipe",40,315,"HT Pro","D/S 1.5m"),
    _p("5405050150-i","D/S Pipe L=1500mm","Pipe",50,413,"HT Pro","D/S 1.5m"),
    _p("5407575150-i","D/S Pipe L=1500mm","Pipe",75,712,"HT Pro","D/S 1.5m"),
    _p("5401111150-i","D/S Pipe L=1500mm","Pipe",110,1157,"HT Pro","D/S 1.5m"),
    _p("5401212150-i","D/S Pipe L=1500mm","Pipe",125,1876,"HT Pro","D/S 1.5m"),
    _p("5401616150-i","D/S Pipe L=1500mm","Pipe",160,2286,"HT Pro","D/S 1.5m"),
    # D/S 2000mm
    _p("5404040200-i","D/S Pipe L=2000mm","Pipe",40,406,"HT Pro","D/S 2m"),
    _p("5405050200-i","D/S Pipe L=2000mm","Pipe",50,531,"HT Pro","D/S 2m"),
    _p("5407575200-i","D/S Pipe L=2000mm","Pipe",75,924,"HT Pro","D/S 2m"),
    _p("5401111200-i","D/S Pipe L=2000mm","Pipe",110,1494,"HT Pro","D/S 2m"),
    _p("5401212200-i","D/S Pipe L=2000mm","Pipe",125,2423,"HT Pro","D/S 2m"),
    _p("5401616200-i","D/S Pipe L=2000mm","Pipe",160,2946,"HT Pro","D/S 2m"),
    # D/S 3000mm
    _p("5404040300-i","D/S Pipe L=3000mm","Pipe",40,588,"HT Pro","D/S 3m"),
    _p("5405050300-i","D/S Pipe L=3000mm","Pipe",50,768,"HT Pro","D/S 3m"),
    _p("5407575300-i","D/S Pipe L=3000mm","Pipe",75,1344,"HT Pro","D/S 3m"),
    _p("5401111300-i","D/S Pipe L=3000mm","Pipe",110,2063,"HT Pro","D/S 3m"),
    _p("5401212300-i","D/S Pipe L=3000mm","Pipe",125,3518,"HT Pro","D/S 3m"),
    _p("5401616300-i","D/S Pipe L=3000mm","Pipe",160,4268,"HT Pro","D/S 3m"),
    # ════════════════════════════════════════
    # HT PRO — BENDS
    # ════════════════════════════════════════
    _p("40040150","Bend 15° DN110","Bend",110,234,"HT Pro","15°"),
    _p("40040350","Bend 30° DN110","Bend",110,239,"HT Pro","30°"),
    _p("40010460-i","Bend 45° DN40","Bend",40,66,"HT Pro","45°"),
    _p("40020460-i","Bend 45° DN50","Bend",50,82,"HT Pro","45°"),
    _p("40030460-i","Bend 45° DN75","Bend",75,128,"HT Pro","45°"),
    _p("40040460-i","Bend 45° DN110","Bend",110,186,"HT Pro","45°"),
    _p("40050460","Bend 45° DN125","Bend",125,768,"HT Pro","45°"),
    _p("40060460","Bend 45° DN160","Bend",160,1013,"HT Pro","45°"),
    _p("40020467","Door Bend 45° DN50","Bend",50,149,"HT Pro","Door 45°"),
    _p("40040467","Door Bend 45° DN110","Bend",110,456,"HT Pro","Door 45°"),
    _p("40040468","Door Bend 45° Right DN110","Bend",110,386,"HT Pro","Door 45°R"),
    _p("40040469","Door Bend 45° Left DN110","Bend",110,382,"HT Pro","Door 45°L"),
    _p("40060457","Door Bend 45° DN160","Bend",160,1751,"HT Pro","Door 45°"),
    _p("40060458","Door Bend 45° Right DN160","Bend",160,1504,"HT Pro","Door 45°R"),
    _p("40060459","Door Bend 45° Left DN160","Bend",160,1626,"HT Pro","Door 45°L"),
    _p("40010860-i","Bend 87.5° DN40","Bend",40,72,"HT Pro","87.5°"),
    _p("40020860-i","Bend 87.5° DN50","Bend",50,77,"HT Pro","87.5°"),
    _p("40030860","Bend 87.5° DN75","Bend",75,155,"HT Pro","87.5°"),
    _p("40040860-i","Bend 87.5° DN110","Bend",110,270,"HT Pro","87.5°"),
    _p("40050860","Bend 87.5° DN125","Bend",125,700,"HT Pro","87.5°"),
    _p("40060860","Bend 87.5° DN160","Bend",160,722,"HT Pro","87.5°"),
    _p("40020867-i","Door Bend 87.5° DN50","Bend",50,221,"HT Pro","Door 87.5°"),
    _p("40030867","Door Bend 87.5° DN75","Bend",75,209,"HT Pro","Door 87.5°"),
    _p("40040867-i","Door Bend 87.5° DN110","Bend",110,517,"HT Pro","Door 87.5°"),
    _p("40060857","Door Bend 87.5° DN160","Bend",160,2429,"HT Pro","Door 87.5°"),
    _p("40040868-i","Door Bend 87.5° Right DN110","Bend",110,553,"HT Pro","Door 87.5°R"),
    _p("40040869","Door Bend 87.5° Left DN110","Bend",110,565,"HT Pro","Door 87.5°L"),
    _p("40060858","Door Bend 87.5° Right DN160","Bend",160,1907,"HT Pro","Door 87.5°R"),
    _p("40060859","Door Bend 87.5° Left DN160","Bend",160,1927,"HT Pro","Door 87.5°L"),
    # ════════════════════════════════════════
    # HT PRO — BRANCHES
    # ════════════════════════════════════════
    _p("40611440","Single Y 45° 40×40","Branch",40,96,"HT Pro","Y 40×40"),
    _p("40622460","Single Y 45° 50×50","Branch",50,139,"HT Pro","Y 50×50"),
    _p("40633460-i","Single Y 45° 75×75","Branch",75,262,"HT Pro","Y 75×75"),
    _p("40644460-i","Single Y 45° 110×110","Branch",110,575,"HT Pro","Y 110×110"),
    _p("40655460","Single Y 45° 125×125","Branch",125,1475,"HT Pro","Y 125×125"),
    _p("40666450","Single Y 45° 160×160","Branch",160,1942,"HT Pro","Y 160×160"),
    _p("40621440","Reducing Y 45° 50×40","Branch",50,143,"HT Pro","Red.Y 50×40"),
    _p("40632460","Reducing Y 45° 75×50","Branch",75,350,"HT Pro","Red.Y 75×50"),
    _p("40642460","Reducing Y 45° 110×50","Branch",110,361,"HT Pro","Red.Y 110×50"),
    _p("40643460","Reducing Y 45° 110×75","Branch",110,578,"HT Pro","Red.Y 110×75"),
    _p("40654460","Reducing Y 45° 125×110","Branch",125,1270,"HT Pro","Red.Y 125×110"),
    _p("40664450","Reducing Y 45° 160×110","Branch",160,1744,"HT Pro","Red.Y 160×110"),
    _p("40611860","Single Tee 90° 40×40","Branch",40,110,"HT Pro","Tee 40×40"),
    _p("40622860","Single Tee 90° 50×50","Branch",50,121,"HT Pro","Tee 50×50"),
    _p("40633860-i","Single Tee 90° 75×75","Branch",75,197,"HT Pro","Tee 75×75"),
    _p("40655860","Single Tee 90° 125×125","Branch",125,607,"HT Pro","Tee 125×125"),
    _p("40666860","Single Tee 90° 160×160","Branch",160,947,"HT Pro","Tee 160×160"),
    _p("40621860","Reducing Tee 90° 50×40","Branch",50,166,"HT Pro","Red.Tee 50×40"),
    _p("40632860","Reducing Tee 90° 75×50","Branch",75,185,"HT Pro","Red.Tee 75×50"),
    _p("40642860","Reducing Tee Short 110×50","Branch",110,392,"HT Pro","Red.Tee 110×50"),
    _p("40652860","Reducing Tee 90° 125×50","Branch",125,614,"HT Pro","Red.Tee 125×50"),
    _p("40662860","Reducing Tee 90° 160×50","Branch",160,1093,"HT Pro","Red.Tee 160×50"),
    _p("40664860","Reducing Tee 90° 160×110","Branch",160,877,"HT Pro","Red.Tee 160×110"),
    _p("40733860-i","Swept Tee 87.5° 75×75","Branch",75,260,"HT Pro","Swept 75×75"),
    _p("40742860-i","Reducing Swept Tee 110×50","Branch",110,436,"HT Pro","Swept 110×50"),
    _p("40743860-i","Reducing Swept Tee 110×75","Branch",110,550,"HT Pro","Swept 110×75"),
    _p("40744860-i","Swept Tee 87.5° 110×110","Branch",110,510,"HT Pro","Swept 110×110"),
    _p("40754860","Reducing Swept Tee 125×110","Branch",125,1080,"HT Pro","Swept 125×110"),
    _p("14040764860-i","Reducing Swept Tee 160×110","Branch",160,2086,"HT Pro","Swept 160×110"),
    _p("41044860","Double Swept Tee 110×110×110","Branch",110,2023,"HT Pro","Dbl Swept"),
    _p("41054850-i","Double Branch 87.5/90° 125×110×110","Branch",125,2241,"HT Pro","Dbl Branch 125"),
    _p("41064850","Double Branch 87.5/90° 160×110×110","Branch",160,2754,"HT Pro","Dbl Branch 160"),
    _p("40733867-i","Door Swept Tee 75×75","Branch",75,460,"HT Pro","Door Swept 75×75"),
    _p("40742867-i","Door Swept Tee 110×50","Branch",110,616,"HT Pro","Door Swept 110×50"),
    _p("40743867-i","Door Swept Tee 110×75","Branch",110,732,"HT Pro","Door Swept 110×75"),
    _p("40744867-i","Door Swept Tee 110×110","Branch",110,754,"HT Pro","Door Swept 110×110"),
    _p("40754867-HM","Door Swept Tee 125×110","Branch",125,2016,"HT Pro","Door Swept 125×110"),
    _p("40664857-i","Door Swept Tee 160×110","Branch",160,2766,"HT Pro","Door Swept 160×110"),
    _p("41044857-HM","Double Swept Door Tee 110×110×110","Branch",110,2232,"HT Pro","Dbl Door Swept"),
    _p("41244850","Corner Branch 87° 110×110×110","Branch",110,1346,"HT Pro","Corner 110"),
    _p("4041254850","Corner Branch 87° 125×110×110","Branch",125,1968,"HT Pro","Corner 125"),
    _p("41264850-i","Red. Corner Branch 160×110×110","Branch",160,2382,"HT Pro","Corner 160"),
    _p("41044450","Double Y 45° 110×110×110","Branch",110,1463,"HT Pro","Dbl Y 110"),
    _p("41042450","Reducing Double Y 45° 110×50×50","Branch",110,1560,"HT Pro","Red.Dbl Y 110×50×50"),
    _p("41062450","Reducing Double Y 45° 160×50×50","Branch",160,1225,"HT Pro","Red.Dbl Y 160×50×50"),
    _p("41044457","Double Y Door 110×110×110","Branch",110,3106,"HT Pro","Dbl Y Door"),
    # ════════════════════════════════════════
    # HT PRO — TRAPS
    # ════════════════════════════════════════
    _p("49540750G-i","P Trap 50mm Water Seal","Trap",110,1132,"HT Pro","P Trap"),
    _p("41840051-i","S Trap Siphon Type","Trap",110,2262,"HT Pro","S Trap"),
    _p("69111750G-i","Nahani Trap 110×75","Trap",110,649,"HT Pro","Nahani Trap"),
    _p("60117051","Multi Floor Trap MFT 7\"","Trap",110,1256,"HT Pro","MFT"),
    _p("14048111060G-i","Multi Floor Trap with Socket","Trap",110,1680,"HT Pro","MFT Socketed"),
    _p("41242850-i","Ht. Riser 2-Inlet 90° 110×50×50","Trap",110,1061,"HT Pro","Riser 2-Inlet 90°"),
    _p("41042860","Riser 2-Inlet 180° 110×50×50","Trap",110,1265,"HT Pro","Riser 2-Inlet 180°"),
    _p("4102042860","Hopper 3-Inlet 110×50×50×50","Trap",110,1380,"HT Pro","Hopper 3-Inlet"),
    _p("4041043280","Riser 50+75mm Inlet 110×75×50","Trap",110,1038,"HT Pro","Riser 75+50mm"),
    _p("69201551G-i","Height Riser L150 for MFT","Trap",110,497,"HT Pro","Riser L150"),
    _p("69203551G-i","Height Riser L350 for MFT","Trap",110,737,"HT Pro","Riser L350"),
    _p("60203651-i","Height Riser Thread Lock Trap","Trap",110,109,"HT Pro","Riser Thread Lock"),
    _p("4049911100G-i","H.A.F.F Stack Single Stack","Trap",110,6600,"HT Pro","HAFF Stack"),
    # ════════════════════════════════════════
    # HT PRO — INSPECTION
    # ════════════════════════════════════════
    _p("49130060-i","Inspection Pipe DN75","Inspection",75,460,"HT Pro","Cleaning Pipe"),
    _p("49140060-i","Inspection Pipe DN110","Inspection",110,566,"HT Pro","Cleaning Pipe"),
    _p("49150060","Inspection Pipe DN125","Inspection",125,1061,"HT Pro","Cleaning Pipe"),
    _p("49160060","Inspection Pipe DN160","Inspection",160,1199,"HT Pro","Cleaning Pipe"),
    # ════════════════════════════════════════
    # HT PRO — COUPLERS / SLEEVES
    # ════════════════════════════════════════
    _p("41710055-i","Coupler DN40","Coupler",40,73,"HT Pro","Coupler"),
    _p("41720055-i","Coupler DN50","Coupler",50,85,"HT Pro","Coupler"),
    _p("41730050-i","Coupler DN75","Coupler",75,115,"HT Pro","Coupler"),
    _p("41740055-i","Coupler DN110","Coupler",110,210,"HT Pro","Coupler"),
    _p("41750065","Coupler DN125","Coupler",125,494,"HT Pro","Coupler"),
    _p("41760055","Coupler DN160","Coupler",160,1054,"HT Pro","Coupler"),
    _p("41720053","Sleeve DN50","Coupler",50,115,"HT Pro","Sleeve"),
    _p("41730053","Sleeve DN75","Coupler",75,128,"HT Pro","Sleeve"),
    _p("41740053","Sleeve DN110","Coupler",110,215,"HT Pro","Sleeve"),
    # ════════════════════════════════════════
    # HT PRO — REDUCERS
    # ════════════════════════════════════════
    _p("42121060","Eccentric Reducer 50×40","Reducer",50,66,"HT Pro","Ecc.Red"),
    _p("42132050","Eccentric Reducer 75×50","Reducer",75,83,"HT Pro","Ecc.Red"),
    _p("42142050","Eccentric Reducer 110×50","Reducer",110,182,"HT Pro","Ecc.Red"),
    _p("42143050-i","Eccentric Reducer 110×75","Reducer",110,215,"HT Pro","Ecc.Red"),
    _p("42154060","Eccentric Reducer 125×110","Reducer",125,421,"HT Pro","Ecc.Red"),
    _p("42164050","Eccentric Reducer 160×110","Reducer",160,653,"HT Pro","Ecc.Red"),
    _p("42134050-i","Eccentric Reverse Reducer 110×75","Reducer",110,510,"HT Pro","Rev.Red"),
    _p("P0500000000040K","Concentric Reducer Bushing 50×40","Reducer",50,192,"HT Pro","Con.Red"),
    _p("P07500000050K","Concentric Reducer Bushing 75×50","Reducer",75,253,"HT Pro","Con.Red"),
    _p("P1100000000050V","Concentric Reducer Bushing 110×50","Reducer",110,328,"HT Pro","Con.Red"),
    _p("P1100000000075V","Concentric Reducer Bushing 110×75","Reducer",110,401,"HT Pro","Con.Red"),
    # ════════════════════════════════════════
    # HT PRO — ACCESSORIES
    # ════════════════════════════════════════
    _p("41610040","End Cap DN40","Accessory",40,35,"HT Pro","End Cap"),
    _p("41620050-i","End Cap DN50","Accessory",50,66,"HT Pro","End Cap"),
    _p("41630050-i","End Cap DN75","Accessory",75,70,"HT Pro","End Cap"),
    _p("41640050-i","End Cap DN110","Accessory",110,115,"HT Pro","End Cap"),
    _p("41650060","End Cap DN125","Accessory",125,251,"HT Pro","End Cap"),
    _p("41660060","End Cap DN160","Accessory",160,344,"HT Pro","End Cap"),
    _p("4042320040","Vent Cowl DN50","Accessory",50,239,"HT Pro","Vent Cowl"),
    _p("4042330040","Vent Cowl DN75","Accessory",75,185,"HT Pro","Vent Cowl"),
    _p("4042340060","Vent Cowl DN110","Accessory",110,194,"HT Pro","Vent Cowl"),
    _p("4042360040","Vent Cowl DN160","Accessory",160,413,"HT Pro","Vent Cowl"),
    _p("5401106650-i","Boss Connector 1-Inlet 110×L660","Accessory",110,1004,"HT Pro","Boss Pipe"),
    _p("540110502-i","Boss Connector Double Branch 110×L660","Accessory",110,1404,"HT Pro","Boss Double"),
    _p("5401105102-i","Boss Connector Corner Branch 110×L660","Accessory",110,1404,"HT Pro","Boss Corner"),
    _p("540110503-i","Boss Connector Triple Branch 110×L660","Accessory",110,1604,"HT Pro","Boss Triple"),
    _p("41540020","WC Connector Straight White","Accessory",110,1558,"HT Pro","WC Conn."),
    _p("41540027","WC Connector with Inspection White","Accessory",110,1776,"HT Pro","WC Door Conn."),
    _p("41542866","WC Bend Back Inspection White","Accessory",110,1559,"HT Pro","WC Bend"),
    _p("41540615","Flange for WC Bend White","Accessory",110,66,"HT Pro","WC Flange"),
    _p("60150331","Aquaslim S.Steel L=330mm","Accessory",0,1288,"HT Pro","Shower Channel"),
    _p("60150339","Aquaslim Full S.Steel 330 Tiles","Accessory",0,1333,"HT Pro","Shower Channel"),
    _p("60150701","Aquaslim S.Steel L=700mm","Accessory",0,1790,"HT Pro","Shower Channel"),
    _p("60150709","Aquaslim Full S.Steel 700 Tiles","Accessory",0,1939,"HT Pro","Shower Channel"),
    _p("60200260","Extension for Square Grating","Accessory",0,82,"HT Pro","Grating Ext"),
    _p("60200263","Extension for Square Grating (alt)","Accessory",0,78,"HT Pro","Grating Ext"),
    _p("47700012","Lubricant 250ml Tin Pack","Accessory",0,198,"HT Pro","Lubricant"),
    # ════════════════════════════════════════
    # HT PRO — CLAMPS
    # ════════════════════════════════════════
    _p("48100040-S","Split Clamp DN40","Clamp",40,126,"HT Pro","Clamp"),
    _p("48100050-S","Split Clamp DN50","Clamp",50,138,"HT Pro","Clamp"),
    _p("48100075-S","Split Clamp DN75","Clamp",75,174,"HT Pro","Clamp"),
    _p("48100011-S","Split Clamp DN110","Clamp",110,210,"HT Pro","Clamp"),
    _p("48100012-S","Split Clamp DN125","Clamp",125,246,"HT Pro","Clamp"),
    _p("48100016-S","Split Clamp DN160","Clamp",160,300,"HT Pro","Clamp"),
    _p("48100020-S","Split Clamp DN200","Clamp",200,354,"HT Pro","Clamp"),

    # ════════════════════════════════════════
    # ULTRA SILENT — PIPES
    # ════════════════════════════════════════
    # S/S 250mm
    _p("5754000025-i","US S/S Pipe L=250mm","Pipe",40,117,"Ultra Silent","S/S 250mm"),
    _p("5755000025-i","US S/S Pipe L=250mm","Pipe",50,152,"Ultra Silent","S/S 250mm"),
    _p("5757500025-i","US S/S Pipe L=250mm","Pipe",75,278,"Ultra Silent","S/S 250mm"),
    _p("5751100025-i","US S/S Pipe L=250mm","Pipe",110,376,"Ultra Silent","S/S 250mm"),
    _p("5751200025-i","US S/S Pipe L=250mm","Pipe",125,470,"Ultra Silent","S/S 250mm"),
    _p("5751600025-i","US S/S Pipe L=250mm","Pipe",160,741,"Ultra Silent","S/S 250mm"),
    # S/S 500mm
    _p("5754000050-i","US S/S Pipe L=500mm","Pipe",40,200,"Ultra Silent","S/S 500mm"),
    _p("5755000050-i","US S/S Pipe L=500mm","Pipe",50,257,"Ultra Silent","S/S 500mm"),
    _p("5757500050-i","US S/S Pipe L=500mm","Pipe",75,483,"Ultra Silent","S/S 500mm"),
    _p("5751100050-i","US S/S Pipe L=500mm","Pipe",110,650,"Ultra Silent","S/S 500mm"),
    _p("5751200050-i","US S/S Pipe L=500mm","Pipe",125,806,"Ultra Silent","S/S 500mm"),
    _p("5751600050-i","US S/S Pipe L=500mm","Pipe",160,1274,"Ultra Silent","S/S 500mm"),
    # S/S 1000mm
    _p("5754000100-i","US S/S Pipe L=1000mm","Pipe",40,366,"Ultra Silent","S/S 1m"),
    _p("5755000100-i","US S/S Pipe L=1000mm","Pipe",50,468,"Ultra Silent","S/S 1m"),
    _p("5757500100-i","US S/S Pipe L=1000mm","Pipe",75,890,"Ultra Silent","S/S 1m"),
    _p("5751100100-i","US S/S Pipe L=1000mm","Pipe",110,1201,"Ultra Silent","S/S 1m"),
    _p("5751200100-i","US S/S Pipe L=1000mm","Pipe",125,1478,"Ultra Silent","S/S 1m"),
    _p("5751600100-i","US S/S Pipe L=1000mm","Pipe",160,2340,"Ultra Silent","S/S 1m"),
    _p("5752000100-i","US S/S Pipe L=1000mm","Pipe",200,2926,"Ultra Silent","S/S 1m"),
    # S/S 1500mm
    _p("5754000150-i","US S/S Pipe L=1500mm","Pipe",40,531,"Ultra Silent","S/S 1.5m"),
    _p("5755000150-i","US S/S Pipe L=1500mm","Pipe",50,680,"Ultra Silent","S/S 1.5m"),
    _p("5757500150-i","US S/S Pipe L=1500mm","Pipe",75,1297,"Ultra Silent","S/S 1.5m"),
    _p("5751100150-i","US S/S Pipe L=1500mm","Pipe",110,1751,"Ultra Silent","S/S 1.5m"),
    _p("5751200150-i","US S/S Pipe L=1500mm","Pipe",125,2390,"Ultra Silent","S/S 1.5m"),
    _p("5751600150-i","US S/S Pipe L=1500mm","Pipe",160,3407,"Ultra Silent","S/S 1.5m"),
    _p("5752000150-i","US S/S Pipe L=1500mm","Pipe",200,4322,"Ultra Silent","S/S 1.5m"),
    # S/S 2000mm
    _p("5754000200-i","US S/S Pipe L=2000mm","Pipe",40,697,"Ultra Silent","S/S 2m"),
    _p("5755000200-i","US S/S Pipe L=2000mm","Pipe",50,890,"Ultra Silent","S/S 2m"),
    _p("5757500200-i","US S/S Pipe L=2000mm","Pipe",75,1705,"Ultra Silent","S/S 2m"),
    _p("5751100200-i","US S/S Pipe L=2000mm","Pipe",110,2300,"Ultra Silent","S/S 2m"),
    _p("5751200200-i","US S/S Pipe L=2000mm","Pipe",125,2821,"Ultra Silent","S/S 2m"),
    _p("5751600200-i","US S/S Pipe L=2000mm","Pipe",160,4472,"Ultra Silent","S/S 2m"),
    _p("5752000200-i","US S/S Pipe L=2000mm","Pipe",200,5718,"Ultra Silent","S/S 2m"),
    # S/S 3000mm
    _p("5753200300-i","US S/S Pipe L=3000mm","Pipe",32,948,"Ultra Silent","S/S 3m"),
    _p("5754000300-i","US S/S Pipe L=3000mm","Pipe",40,1027,"Ultra Silent","S/S 3m"),
    _p("5755000300-i","US S/S Pipe L=3000mm","Pipe",50,1313,"Ultra Silent","S/S 3m"),
    _p("5757500300-i","US S/S Pipe L=3000mm","Pipe",75,2521,"Ultra Silent","S/S 3m"),
    _p("5751100300-i","US S/S Pipe L=3000mm","Pipe",110,2953,"Ultra Silent","S/S 3m"),
    _p("5751200300-i","US S/S Pipe L=3000mm","Pipe",125,4165,"Ultra Silent","S/S 3m"),
    _p("5751600300-i","US S/S Pipe L=3000mm","Pipe",160,6060,"Ultra Silent","S/S 3m"),
    _p("5752000300-i","US S/S Pipe L=3000mm","Pipe",200,8520,"Ultra Silent","S/S 3m"),
    # D/S 500mm
    _p("5754040050-i","US D/S Pipe L=500mm","Pipe",40,234,"Ultra Silent","D/S 500mm"),
    _p("5755050050-i","US D/S Pipe L=500mm","Pipe",50,302,"Ultra Silent","D/S 500mm"),
    _p("5757575050-i","US D/S Pipe L=500mm","Pipe",75,556,"Ultra Silent","D/S 500mm"),
    _p("5751111050-i","US D/S Pipe L=500mm","Pipe",110,751,"Ultra Silent","D/S 500mm"),
    _p("5751212050-i","US D/S Pipe L=500mm","Pipe",125,947,"Ultra Silent","D/S 500mm"),
    _p("5751616050-i","US D/S Pipe L=500mm","Pipe",160,1482,"Ultra Silent","D/S 500mm"),
    # D/S 1000mm
    _p("5754040100-i","US D/S Pipe L=1000mm","Pipe",40,400,"Ultra Silent","D/S 1m"),
    _p("5755050100-i","US D/S Pipe L=1000mm","Pipe",50,514,"Ultra Silent","D/S 1m"),
    _p("5757575100-i","US D/S Pipe L=1000mm","Pipe",75,964,"Ultra Silent","D/S 1m"),
    _p("5751111100-i","US D/S Pipe L=1000mm","Pipe",110,1301,"Ultra Silent","D/S 1m"),
    _p("5751212100-i","US D/S Pipe L=1000mm","Pipe",125,1613,"Ultra Silent","D/S 1m"),
    _p("5751616100-i","US D/S Pipe L=1000mm","Pipe",160,2547,"Ultra Silent","D/S 1m"),
    _p("5752020100-i","US D/S Pipe L=1000mm","Pipe",200,3059,"Ultra Silent","D/S 1m"),
    # D/S 1500mm
    _p("5754040150-i","US D/S Pipe L=1500mm","Pipe",40,564,"Ultra Silent","D/S 1.5m"),
    _p("5755050150-i","US D/S Pipe L=1500mm","Pipe",50,724,"Ultra Silent","D/S 1.5m"),
    _p("5757575150-i","US D/S Pipe L=1500mm","Pipe",75,1371,"Ultra Silent","D/S 1.5m"),
    _p("5751111150-i","US D/S Pipe L=1500mm","Pipe",110,1850,"Ultra Silent","D/S 1.5m"),
    _p("5751212150-i","US D/S Pipe L=1500mm","Pipe",125,2470,"Ultra Silent","D/S 1.5m"),
    _p("5751616150-i","US D/S Pipe L=1500mm","Pipe",160,3614,"Ultra Silent","D/S 1.5m"),
    _p("5752020150-i","US D/S Pipe L=1500mm","Pipe",200,4455,"Ultra Silent","D/S 1.5m"),
    # D/S 2000mm
    _p("5754040200-i","US D/S Pipe L=2000mm","Pipe",40,730,"Ultra Silent","D/S 2m"),
    _p("5755050200-i","US D/S Pipe L=2000mm","Pipe",50,935,"Ultra Silent","D/S 2m"),
    _p("5757575200-i","US D/S Pipe L=2000mm","Pipe",75,1779,"Ultra Silent","D/S 2m"),
    _p("5751111200-i","US D/S Pipe L=2000mm","Pipe",110,2401,"Ultra Silent","D/S 2m"),
    _p("5751212200-i","US D/S Pipe L=2000mm","Pipe",125,2957,"Ultra Silent","D/S 2m"),
    _p("5751616200-i","US D/S Pipe L=2000mm","Pipe",160,4680,"Ultra Silent","D/S 2m"),
    _p("5752020200-i","US D/S Pipe L=2000mm","Pipe",200,5851,"Ultra Silent","D/S 2m"),
    # D/S 3000mm
    _p("5754040300-i","US D/S Pipe L=3000mm","Pipe",40,1061,"Ultra Silent","D/S 3m"),
    _p("5755050300-i","US D/S Pipe L=3000mm","Pipe",50,1358,"Ultra Silent","D/S 3m"),
    _p("5757575300-i","US D/S Pipe L=3000mm","Pipe",75,2593,"Ultra Silent","D/S 3m"),
    _p("5751111300-i","US D/S Pipe L=3000mm","Pipe",110,3110,"Ultra Silent","D/S 3m"),
    _p("5751212300-i","US D/S Pipe L=3000mm","Pipe",125,4300,"Ultra Silent","D/S 3m"),
    _p("5751616300-i","US D/S Pipe L=3000mm","Pipe",160,6812,"Ultra Silent","D/S 3m"),
    _p("5752020300-i","US D/S Pipe L=3000mm","Pipe",200,8653,"Ultra Silent","D/S 3m"),
    # ════════════════════════════════════════
    # ULTRA SILENT — BENDS
    # ════════════════════════════════════════
    _p("7070000170","US Bend 15° DN32","Bend",32,61,"Ultra Silent","15°"),
    _p("7070010170","US Bend 15° DN40","Bend",40,82,"Ultra Silent","15°"),
    _p("7070020170","US Bend 15° DN50","Bend",50,102,"Ultra Silent","15°"),
    _p("7070030170","US Bend 15° DN75","Bend",75,184,"Ultra Silent","15°"),
    _p("7070040170","US Bend 15° DN110","Bend",110,698,"Ultra Silent","15°"),
    _p("7070050170","US Bend 15° DN125","Bend",125,694,"Ultra Silent","15°"),
    _p("7070060170","US Bend 15° DN160","Bend",160,1357,"Ultra Silent","15°"),
    _p("7070000370","US Bend 30° DN32","Bend",32,82,"Ultra Silent","30°"),
    _p("7070010370","US Bend 30° DN40","Bend",40,102,"Ultra Silent","30°"),
    _p("7070020370","US Bend 30° DN50","Bend",50,102,"Ultra Silent","30°"),
    _p("7070030370","US Bend 30° DN75","Bend",75,205,"Ultra Silent","30°"),
    _p("7070040370","US Bend 30° DN110","Bend",110,727,"Ultra Silent","30°"),
    _p("7070050370","US Bend 30° DN125","Bend",125,838,"Ultra Silent","30°"),
    _p("7070060370","US Bend 30° DN160","Bend",160,1330,"Ultra Silent","30°"),
    _p("7070000470","US Bend 45° DN32","Bend",32,61,"Ultra Silent","45°"),
    _p("7070010470","US Bend 45° DN40","Bend",40,82,"Ultra Silent","45°"),
    _p("7070020470","US Bend 45° DN50","Bend",50,122,"Ultra Silent","45°"),
    _p("7070030470-i","US Bend 45° DN75","Bend",75,205,"Ultra Silent","45°"),
    _p("7070040470-i","US Bend 45° DN110","Bend",110,641,"Ultra Silent","45°"),
    _p("7070050470","US Bend 45° DN125","Bend",125,899,"Ultra Silent","45°"),
    _p("7070060470","US Bend 45° DN160","Bend",160,1202,"Ultra Silent","45°"),
    _p("7070080470","US Bend 45° DN200","Bend",200,3493,"Ultra Silent","45°"),
    _p("7070000670","US Bend 67.5° DN32","Bend",32,61,"Ultra Silent","67.5°"),
    _p("7070010670","US Bend 67.5° DN40","Bend",40,61,"Ultra Silent","67.5°"),
    _p("7070020670","US Bend 67.5° DN50","Bend",50,82,"Ultra Silent","67.5°"),
    _p("7070030670","US Bend 67.5° DN75","Bend",75,122,"Ultra Silent","67.5°"),
    _p("7070040670","US Bend 67.5° DN110","Bend",110,632,"Ultra Silent","67.5°"),
    _p("7070050670","US Bend 67.5° DN125","Bend",125,643,"Ultra Silent","67.5°"),
    _p("7070000870","US Bend 87.5° DN32","Bend",32,61,"Ultra Silent","87.5°"),
    _p("7070010870","US Bend 87.5° DN40","Bend",40,102,"Ultra Silent","87.5°"),
    _p("7070020870-i","US Bend 87.5° DN50","Bend",50,143,"Ultra Silent","87.5°"),
    _p("7070030870","US Bend 87.5° DN75","Bend",75,224,"Ultra Silent","87.5°"),
    _p("7070040870","US Bend 87.5° DN110","Bend",110,697,"Ultra Silent","87.5°"),
    _p("7070050870","US Bend 87.5° DN125","Bend",125,1062,"Ultra Silent","87.5°"),
    _p("7070060870","US Bend 87.5° DN160","Bend",160,2122,"Ultra Silent","87.5°"),
    _p("7070040877-i","US Door Bend 87.5° DN110","Bend",110,1054,"Ultra Silent","Door 87.5°"),
    # ════════════════════════════════════════
    # ULTRA SILENT — BRANCHES
    # ════════════════════════════════════════
    _p("7070600470","US Wye 45° 32×32","Branch",32,194,"Ultra Silent","Y 32×32"),
    _p("7070611470","US Wye 45° 40×40","Branch",40,274,"Ultra Silent","Y 40×40"),
    _p("7070621470","US Wye 45° 50×40","Branch",50,352,"Ultra Silent","Y 50×40"),
    _p("7070622470","US Wye 45° 50×50","Branch",50,390,"Ultra Silent","Y 50×50"),
    _p("7070632470","US Wye 45° 75×50","Branch",75,523,"Ultra Silent","Y 75×50"),
    _p("7070633470","US Wye 45° 75×75","Branch",75,698,"Ultra Silent","Y 75×75"),
    _p("7070642470","US Wye 45° 110×50","Branch",110,640,"Ultra Silent","Y 110×50"),
    _p("7070643470-i","US Wye 45° 110×75","Branch",110,959,"Ultra Silent","Y 110×75"),
    _p("7070644470-i","US Wye 45° 110×110","Branch",110,1183,"Ultra Silent","Y 110×110"),
    _p("7070654470","US Wye 45° 125×110","Branch",125,1489,"Ultra Silent","Y 125×110"),
    _p("7070655470","US Wye 45° 125×125","Branch",125,1754,"Ultra Silent","Y 125×125"),
    _p("7070664470","US Wye 45° 160×110","Branch",160,2672,"Ultra Silent","Y 160×110"),
    _p("7070666470","US Wye 45° 160×160","Branch",160,3653,"Ultra Silent","Y 160×160"),
    _p("7070686470","US Wye 45° 200×160","Branch",200,6776,"Ultra Silent","Y 200×160"),
    _p("7070688470","US Wye 45° 200×200","Branch",200,7529,"Ultra Silent","Y 200×200"),
    _p("7070611870","US Tee 90° 40×40","Branch",40,312,"Ultra Silent","Tee 40×40"),
    _p("7070621870","US Tee 90° 50×40","Branch",50,352,"Ultra Silent","Tee 50×40"),
    _p("7070622870","US Tee 90° 50×50","Branch",50,352,"Ultra Silent","Tee 50×50"),
    _p("7070632870","US Tee 90° 75×50","Branch",75,553,"Ultra Silent","Tee 75×50"),
    _p("7070633870","US Tee 90° 75×75","Branch",75,781,"Ultra Silent","Tee 75×75"),
    _p("7070642870","US Tee 90° 110×50","Branch",110,898,"Ultra Silent","Tee 110×50"),
    _p("7070666870","US Tee 90° 160×160","Branch",160,3122,"Ultra Silent","Tee 160×160"),
    _p("7070744870","US Swept Tee 87.5° 110×110","Branch",110,1081,"Ultra Silent","Swept 110×110"),
    _p("7070743870","US Swept Tee 87.5° 110×75","Branch",110,938,"Ultra Silent","Swept 110×75"),
    _p("7070754870","US Swept Tee 87.5° 125×110","Branch",125,1062,"Ultra Silent","Swept 125×110"),
    _p("7070764870","US Swept Tee 87.5° 160×110","Branch",160,2366,"Ultra Silent","Swept 160×110"),
    _p("7070744877","US Door Swept Tee 110×110","Branch",110,1342,"Ultra Silent","Door Swept 110×110"),
    _p("7070754877-i","US Door Swept Tee 125×110","Branch",125,1740,"Ultra Silent","Door Swept 125×110"),
    _p("7070764877-i","US Door Swept Tee 160×110","Branch",160,2752,"Ultra Silent","Door Swept 160×110"),
    _p("7071044870","US Double Swept Tee 110×110×110","Branch",110,2970,"Ultra Silent","Dbl Swept 110"),
    _p("7071244870","US Corner Branch 110×110×110","Branch",110,1285,"Ultra Silent","Corner 110"),
    _p("7071254870","US Corner Branch 125×110×110","Branch",125,1489,"Ultra Silent","Corner 125"),
    _p("7071264870-i","US Corner Branch 160×110×110","Branch",160,2582,"Ultra Silent","Corner 160"),
    _p("7071042670","US Double Branch 67.5° 110×50×50","Branch",110,1105,"Ultra Silent","Dbl 67.5° 110×50×50"),
    _p("7071044670","US Double Branch 67.5° 110×110×110","Branch",110,1518,"Ultra Silent","Dbl 67.5° 110×110×110"),
    _p("7071054870-i","US Double Branch 87.5/90° 125×110×110","Branch",125,2441,"Ultra Silent","Dbl Branch 125"),
    # ════════════════════════════════════════
    # ULTRA SILENT — TRAPS
    # ════════════════════════════════════════
    _p("49540750B-i","US P Trap 50mm Water Seal","Trap",110,1435,"Ultra Silent","P Trap"),
    _p("7071840070-i","US S Trap 110mm","Trap",110,2870,"Ultra Silent","S Trap"),
    _p("60117060","US Multi Floor Trap W/O Ring","Trap",110,1496,"Ultra Silent","MFT W/O Ring"),
    _p("S11050505075-i","US Multi Floor Trap With Ring","Trap",110,1778,"Ultra Silent","MFT With Ring"),
    _p("17078111070-B","US Multi Floor Trap with Socket","Trap",110,1835,"Ultra Silent","MFT+Socket"),
    _p("69111750B-i","US Nahani Trap","Trap",110,781,"Ultra Silent","Nahani Trap"),
    _p("7079911100B-i","US H.A.F.F Stack","Trap",110,9000,"Ultra Silent","HAFF Stack"),
    _p("70114500","SmartLock Trap 140/50 Single Discharge","Trap",50,1243,"Ultra Silent","SmartLock Single"),
    _p("70124599","SmartLock Trap 245/50 Single Discharge","Trap",50,1499,"Ultra Silent","SmartLock Single"),
    _p("70114590","SmartLock Trap 140/40/50 Multi Discharge","Trap",50,1820,"Ultra Silent","SmartLock Multi"),
    _p("70124590","SmartLock Trap 245/40/50 Multi Discharge","Trap",50,2047,"Ultra Silent","SmartLock Multi"),
    _p("70140760","Collector 70/40 Single W/O Trap","Trap",40,760,"Ultra Silent","Collector 70/40"),
    # US Height Riser for P-Trap
    _p("7071042877-i","US Double Branch 180° 110×50×50","Trap",110,1289,"Ultra Silent","Riser 180° 50mm"),
    _p("7071242877-i","US Corner Branch 90° 110×50×50","Trap",110,1459,"Ultra Silent","Corner Riser 90°"),
    _p("70712","US Hopper 3-Inlet 110×50×50×50","Trap",110,1674,"Ultra Silent","Hopper 3-Inlet"),
    _p("70713-i","US Hopper 3-Inlet 110×75×75×75","Trap",110,1958,"Ultra Silent","Hopper 3-Inlet 75"),
    _p("7071043870-i","US Double Branch 180° 110×75×75","Trap",110,1778,"Ultra Silent","Riser 180° 75mm"),
    _p("7071243870-HM","US Corner Branch 90° 110×75×75","Trap",110,1778,"Ultra Silent","Corner Riser 75mm"),
    # US Height Riser for MFT
    _p("69201551B-i","US Height Riser L150 for MFT","Trap",110,650,"Ultra Silent","Riser L150"),
    _p("69203551B-i","US Height Riser L350 for MFT","Trap",110,977,"Ultra Silent","Riser L350"),
    _p("60203651-i","Height Riser for SmartLock Trap","Trap",110,109,"Ultra Silent","Riser SmartLock"),
    # ════════════════════════════════════════
    # ULTRA SILENT — INSPECTION
    # ════════════════════════════════════════
    _p("7079120070","US Inspection Pipe DN50","Inspection",50,349,"Ultra Silent","Inspection"),
    _p("7079130070","US Inspection Pipe DN75","Inspection",75,821,"Ultra Silent","Inspection"),
    _p("7079140070","US Inspection Pipe DN110","Inspection",110,1530,"Ultra Silent","Inspection"),
    _p("7079150070","US Inspection Pipe DN125","Inspection",125,1775,"Ultra Silent","Inspection"),
    _p("7079160070","US Inspection Pipe DN160","Inspection",160,1859,"Ultra Silent","Inspection"),
    _p("7079180070","US Inspection Pipe DN200","Inspection",200,5498,"Ultra Silent","Inspection"),
    # ════════════════════════════════════════
    # ULTRA SILENT — COUPLERS / SLEEVES
    # ════════════════════════════════════════
    _p("7071700270","US Double Socket DN32","Coupler",32,157,"Ultra Silent","Coupler"),
    _p("7071710270","US Double Socket DN40","Coupler",40,193,"Ultra Silent","Coupler"),
    _p("7071720275","US One Way Socket DN50","Coupler",50,194,"Ultra Silent","Coupler"),
    _p("7071730275-i","US One Way Socket DN75","Coupler",75,352,"Ultra Silent","Coupler"),
    _p("7071740275-i","US One Way Socket DN110","Coupler",110,640,"Ultra Silent","Coupler"),
    _p("7071750275","US One Way Socket DN125","Coupler",125,878,"Ultra Silent","Coupler"),
    _p("7071760275","US One Way Socket DN160","Coupler",160,1717,"Ultra Silent","Coupler"),
    _p("7071780275","US One Way Socket DN200","Coupler",200,2594,"Ultra Silent","Coupler"),
    _p("7071710070","US Sleeve DN40","Coupler",40,349,"Ultra Silent","Sleeve"),
    _p("7071720070","US Sleeve DN50","Coupler",50,821,"Ultra Silent","Sleeve"),
    _p("7071730070","US Sleeve DN75","Coupler",75,898,"Ultra Silent","Sleeve"),
    _p("7071740070","US Sleeve DN110","Coupler",110,1775,"Ultra Silent","Sleeve"),
    _p("7071750070","US Sleeve DN125","Coupler",125,1913,"Ultra Silent","Sleeve"),
    _p("7071760070","US Sleeve DN160","Coupler",160,5498,"Ultra Silent","Sleeve"),
    _p("41740060","Double Socket Cast Iron + Seal DN110","Coupler",110,1019,"Ultra Silent","CI Socket"),
    _p("41760051","Double Socket Cast Iron + Seal DN160","Coupler",160,1192,"Ultra Silent","CI Socket"),
    # ════════════════════════════════════════
    # ULTRA SILENT — REDUCERS
    # ════════════════════════════════════════
    _p("7072110070","US Reducer 40×32","Reducer",40,386,"Ultra Silent","Red 40×32"),
    _p("7072120070","US Reducer 50×32","Reducer",50,391,"Ultra Silent","Red 50×32"),
    _p("7072121070","US Reducer 50×40","Reducer",50,398,"Ultra Silent","Red 50×40"),
    _p("7072132070","US Reducer 75×50","Reducer",75,426,"Ultra Silent","Red 75×50"),
    _p("7072142070","US Reducer 110×50","Reducer",110,469,"Ultra Silent","Red 110×50"),
    _p("7072143070","US Reducer 110×75","Reducer",110,491,"Ultra Silent","Red 110×75"),
    _p("7072154070","US Reducer 125×110","Reducer",125,613,"Ultra Silent","Red 125×110"),
    _p("7072164070","US Reducer 160×110","Reducer",160,1021,"Ultra Silent","Red 160×110"),
    _p("7072165070","US Reducer 160×125","Reducer",160,1163,"Ultra Silent","Red 160×125"),
    _p("7072186070","US Reducer 200×160","Reducer",200,2327,"Ultra Silent","Red 200×160"),
    _p("7072134070-i","US Reverse Reducer 110×75","Reducer",110,510,"Ultra Silent","Rev.Red 110×75"),
    # ════════════════════════════════════════
    # ULTRA SILENT — ACCESSORIES
    # ════════════════════════════════════════
    _p("7071610070","US End Cap DN40","Accessory",40,40,"Ultra Silent","End Cap"),
    _p("7071620070-i","US End Cap DN50","Accessory",50,78,"Ultra Silent","End Cap"),
    _p("7071630070","US End Cap DN75","Accessory",75,157,"Ultra Silent","End Cap"),
    _p("7071640070-i","US End Cap DN110","Accessory",110,352,"Ultra Silent","End Cap"),
    _p("7071650070","US End Cap DN125","Accessory",125,757,"Ultra Silent","End Cap"),
    _p("7071660070","US End Cap DN160","Accessory",160,768,"Ultra Silent","End Cap"),
    _p("7071680070","US End Cap DN200","Accessory",200,1616,"Ultra Silent","End Cap"),
    _p("7072330000","US Lock Seal DN75","Accessory",75,434,"Ultra Silent","Lock Seal"),
    _p("7072340000","US Lock Seal DN110","Accessory",110,965,"Ultra Silent","Lock Seal"),
    _p("7072350000","US Lock Seal DN125","Accessory",125,1115,"Ultra Silent","Lock Seal"),
    _p("7072360000","US Lock Seal DN160","Accessory",160,1372,"Ultra Silent","Lock Seal"),
    _p("7072380000","US Lock Seal DN200","Accessory",200,6918,"Ultra Silent","Lock Seal"),
    _p("7078004000","US End Lock DN110","Accessory",110,1518,"Ultra Silent","End Lock"),
    _p("7078005000","US End Lock DN125","Accessory",125,1663,"Ultra Silent","End Lock"),
    _p("7078006000","US End Lock DN160","Accessory",160,1702,"Ultra Silent","End Lock"),
    _p("7078008000","US End Lock DN200","Accessory",200,9194,"Ultra Silent","End Lock"),
    _p("42320040","US Vent Cowl DN50","Accessory",50,239,"Ultra Silent","Vent Cowl"),
    _p("42330040","US Vent Cowl DN75","Accessory",75,185,"Ultra Silent","Vent Cowl"),
    _p("42340060","US Vent Cowl DN110","Accessory",110,194,"Ultra Silent","Vent Cowl"),
    _p("42360040","US Vent Cowl DN160","Accessory",160,413,"Ultra Silent","Vent Cowl"),
    _p("41540020","WC Connector Straight White","Accessory",110,1558,"Ultra Silent","WC Conn."),
    _p("41540027","WC Connector with Inspection White","Accessory",110,1776,"Ultra Silent","WC Door Conn."),
    _p("41542866","WC Bend Back Inspection White","Accessory",110,1559,"Ultra Silent","WC Bend"),
    _p("7074010970","USSW Technical Bend 46mm","Accessory",46,193,"Ultra Silent","Tech Bend"),
    _p("7074021970","USSW Technical Bend 50mm","Accessory",50,215,"Ultra Silent","Tech Bend"),
    _p("7074011970","USSW Technical Bend 46mm (alt)","Accessory",46,193,"Ultra Silent","Tech Bend"),
    _p("7074022970","USSW Technical Bend 50mm (alt)","Accessory",50,193,"Ultra Silent","Tech Bend"),
    _p("7074021971","Long USSW Technical Bend 50mm","Accessory",50,302,"Ultra Silent","Long Tech Bend"),
    _p("7074011971","Long USSW Technical Bend 46mm","Accessory",46,302,"Ultra Silent","Long Tech Bend"),
    _p("T047T000000000","Rubber Gasket US/USSW 46mm (A)","Accessory",46,217,"Ultra Silent","Rubber Gasket"),
    _p("T046T000000000","Rubber Gasket US/USSW 46mm (B)","Accessory",46,217,"Ultra Silent","Rubber Gasket"),
    _p("T050T000000032","Rubber Gasket US/USSW 50mm (A)","Accessory",50,308,"Ultra Silent","Rubber Gasket"),
    _p("T050T000000040","Rubber Gasket US/USSW 50mm (B)","Accessory",50,308,"Ultra Silent","Rubber Gasket"),
    _p("47700012","Lubricant 250ml Tin Pack","Accessory",0,198,"Ultra Silent","Lubricant"),
    # ════════════════════════════════════════
    # ULTRA SILENT — CLAMPS
    # ════════════════════════════════════════
    _p("7890004070-S","US HD Split Clamp DN40","Clamp",40,228,"Ultra Silent","HD Clamp"),
    _p("7890005070-S","US HD Split Clamp DN50","Clamp",50,258,"Ultra Silent","HD Clamp"),
    _p("7890007570-S","US HD Split Clamp DN75","Clamp",75,312,"Ultra Silent","HD Clamp"),
    _p("7890011070-S","US HD Split Clamp DN110","Clamp",110,408,"Ultra Silent","HD Clamp"),
    _p("7890012570-S","US HD Split Clamp DN125","Clamp",125,432,"Ultra Silent","HD Clamp"),
    _p("7890016070-S","US HD Split Clamp DN160","Clamp",160,540,"Ultra Silent","HD Clamp"),
    _p("7890020070-S","US HD Split Clamp DN200","Clamp",200,630,"Ultra Silent","HD Clamp"),
]

# ── CONSTANTS ──
SH_CODES  = ["NA"] + [f"SH{i:02d}" for i in range(1,51)] + [f"K{i:02d}" for i in range(1,51)]
QUICK_SH  = ["NA"] + [f"SH{i:02d}" for i in range(1,11)] + [f"K{i:02d}" for i in range(1,6)]
CATS      = ["Pipe","Bend","Branch","Trap","Coupler","Reducer","Inspection","Clamp","Accessory"]
CAT_ICONS = {"Pipe":"▭","Bend":"↩","Branch":"⑂","Trap":"⊔","Coupler":"○","Reducer":"◁","Inspection":"◎","Clamp":"⊓","Accessory":"⚙"}
DN_LIST   = [32,40,46,50,75,110,125,160,200]
DN_COLORS = {32:"#7C3AED",40:"#1D4ED8",46:"#0E7490",50:"#0E7490",75:"#065F46",110:"#B45309",125:"#B91C1C",160:"#9D174D",200:"#3730A3"}

# ── SESSION STATE ──
for k,v in {"boq":[],"global_sh":"NA","line":"HT Pro","dn_f":None,"cat_f":None,"srch":"","proj":"","g_disc":0,"boq_fsh":"ALL","copy_from":"","copy_to":"SH01"}.items():
    if k not in st.session_state: st.session_state[k]=v

# ── HELPERS ──
def sh_colors(c):
    if c=="NA": return "#64748B","#F1F5F9","#94A3B8"
    if c.startswith("SH"): return "#1D4ED8","#DBEAFE","#3B82F6"
    return "#065F46","#D1FAE5","#10B981"

def fmt(n): return f"₹{int(round(n)):,}"
def eff_disc(b): return b["disc"] if b["disc"] is not None else st.session_state.g_disc
def net_rate(b): return b["price"]*(1-eff_disc(b)/100)
def amt(b): return b["qty"]*net_rate(b)
def grand(): return sum(amt(b) for b in st.session_state.boq)

def get_filtered():
    ps=[p for p in PRODUCTS if p["line"]==st.session_state.line]
    if st.session_state.dn_f: ps=[p for p in ps if p["dn"]==st.session_state.dn_f]
    if st.session_state.cat_f: ps=[p for p in ps if p["type"]==st.session_state.cat_f]
    if st.session_state.srch:
        q=st.session_state.srch.lower()
        ps=[p for p in ps if q in p["desc"].lower() or q in p["code"].lower() or q in p.get("sub","").lower()]
    return ps

def get_dn_counts():
    m={}
    for p in PRODUCTS:
        if p["line"]==st.session_state.line and p["dn"]>0:
            m[p["dn"]]=m.get(p["dn"],0)+1
    return m

def sh_label(c):
    if c=="NA": return "NA — Unassigned"
    if c.startswith("SH"): return f"{c} — Shaft {int(c[2:])}"
    return f"{c} — Kitchen/Flat {int(c[1:])}"

def add_to_boq(prod, qty):
    sh=st.session_state.global_sh
    for item in st.session_state.boq:
        if item["code"]==prod["code"] and item["shaft"]==sh:
            item["qty"]+=qty; return
    st.session_state.boq.append({**prod,"qty":qty,"disc":None,"shaft":sh,"_id":str(uuid.uuid4())[:8]})

def boq_sh_counts():
    m={}
    for b in st.session_state.boq: m[b.get("shaft","NA")]=m.get(b.get("shaft","NA"),0)+1
    return m

def agg_boq():
    m={}
    for b in st.session_state.boq:
        k=b["code"]
        if k not in m: m[k]={**b,"qty":b["qty"],"shafts":[b.get("shaft","NA")]}
        else:
            m[k]["qty"]+=b["qty"]
            sh=b.get("shaft","NA")
            if sh not in m[k]["shafts"]: m[k]["shafts"].append(sh)
    return list(m.values())

def shaft_groups():
    m={}
    for b in st.session_state.boq: m.setdefault(b.get("shaft","NA"),[]).append(b)
    return m

# ── EXCEL EXPORT ──
def export_excel():
    wb=openpyxl.Workbook()
    boq=st.session_state.boq
    INR,INT,PCT='₹#,##0.00','#,##0','0.00"%"'
    hdr_fill=PatternFill("solid",fgColor="0F172A")
    hdr_font=Font(bold=True,color="FFFFFF",name="Calibri",size=10)
    hdr_align=Alignment(horizontal="center",vertical="center")
    alt_fill=PatternFill("solid",fgColor="F8FAFC")
    tot_fill=PatternFill("solid",fgColor="1E3A5F")
    tot_font=Font(bold=True,color="FBBF24",name="Calibri",size=12)
    row_border=Border(bottom=Side(style="thin",color="E2E8F0"))

    def write_header(ws,headers,widths,fill=None):
        f=fill or hdr_fill
        for c,(h,w) in enumerate(zip(headers,widths),1):
            cell=ws.cell(1,c,h); cell.font=hdr_font; cell.fill=f; cell.alignment=hdr_align
            ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width=w
        ws.row_dimensions[1].height=22; ws.freeze_panes="A2"

    def sa(cell,align="left"):
        cell.alignment=Alignment(horizontal=align,vertical="center")

    # ── Full BOQ ──
    ws=wb.active; ws.title="Full BOQ"; ws.sheet_view.showGridLines=False
    write_header(ws,["Sr.No","Item Code","Description","Type","DN","Unit","Total Qty","List Price (₹)","Disc %","Net Rate (₹)","Amount (₹)","Shafts / Location","Line"],
                 [6,22,42,14,6,5,9,14,7,14,14,28,12])
    agg=agg_boq()
    for i,r in enumerate(agg,2):
        ws.cell(i,1,i-1); sa(ws.cell(i,1),"center")
        ws.cell(i,2,r["code"]); ws.cell(i,3,r["desc"]); ws.cell(i,4,r["sub"])
        ws.cell(i,5,r["dn"] if r["dn"] else "-"); sa(ws.cell(i,5),"center")
        ws.cell(i,6,"Nos"); sa(ws.cell(i,6),"center")
        c_qty=ws.cell(i,7,r["qty"]); c_qty.number_format=INT; c_qty.font=Font(bold=True,color="1D4ED8",name="Calibri",size=11); sa(c_qty,"center")
        c_lp=ws.cell(i,8,r["price"]); c_lp.number_format=INR; sa(c_lp,"right")
        c_d=ws.cell(i,9,eff_disc(r)); c_d.number_format=PCT; sa(c_d,"center")
        c_nr=ws.cell(i,10,f"=H{i}*(1-I{i}/100)"); c_nr.number_format=INR; c_nr.font=Font(color="10B981",name="Calibri",size=10,italic=True); sa(c_nr,"right")
        c_amt=ws.cell(i,11,f"=G{i}*J{i}"); c_amt.number_format=INR; sa(c_amt,"right")
        c_sh=ws.cell(i,12,", ".join(sorted(r["shafts"]))); c_sh.fill=PatternFill("solid",fgColor="DBEAFE"); c_sh.font=Font(color="1D4ED8",name="Calibri",size=9)
        ws.cell(i,13,r["line"])
        for c in range(1,14): ws.cell(i,c).border=row_border
        if i%2==0:
            for c in [1,2,3,4,5,6,8,9,10,11,13]: ws.cell(i,c).fill=alt_fill
    last=len(agg)+1; gt=last+2
    ws.cell(gt,3,"GRAND TOTAL — ALL SHAFTS COMBINED").font=Font(bold=True,name="Calibri",size=11)
    c_gq=ws.cell(gt,7,f"=SUM(G2:G{last})"); c_gq.number_format=INT; c_gq.font=Font(bold=True,color="1D4ED8",name="Calibri",size=12); sa(c_gq,"center")
    c_ga=ws.cell(gt,11,f"=SUM(K2:K{last})"); c_ga.number_format=INR; c_ga.font=tot_font; c_ga.fill=tot_fill; sa(c_ga,"right")
    ws.cell(gt,12,f"{len(boq)} entries · {len(set(b.get('shaft','NA') for b in boq))} locations").font=Font(italic=True,color="64748B",name="Calibri",size=9)

    # ── Detail BOQ ──
    ws_d=wb.create_sheet("Detail BOQ"); ws_d.sheet_view.showGridLines=False
    write_header(ws_d,["Sr.No","Location","Item Code","Description","Type","DN","Unit","Qty","List Price (₹)","Disc %","Net Rate (₹)","Amount (₹)","Line"],
                 [6,10,22,42,14,6,5,6,14,7,14,14,12])
    for i,b in enumerate(boq,2):
        fg=("DBEAFE" if b.get("shaft","NA").startswith("SH") else "D1FAE5" if b.get("shaft","NA").startswith("K") else "F1F5F9")
        fc=("1D4ED8" if b.get("shaft","NA").startswith("SH") else "065F46" if b.get("shaft","NA").startswith("K") else "64748B")
        ws_d.cell(i,1,i-1); sa(ws_d.cell(i,1),"center")
        c_loc=ws_d.cell(i,2,b.get("shaft","NA")); c_loc.fill=PatternFill("solid",fgColor=fg); c_loc.font=Font(bold=True,name="Calibri",size=10,color=fc); sa(c_loc,"center")
        ws_d.cell(i,3,b["code"]); ws_d.cell(i,4,b["desc"]); ws_d.cell(i,5,b["sub"])
        ws_d.cell(i,6,b["dn"] if b["dn"] else "-"); sa(ws_d.cell(i,6),"center")
        ws_d.cell(i,7,"Nos"); sa(ws_d.cell(i,7),"center")
        c_q2=ws_d.cell(i,8,b["qty"]); c_q2.number_format=INT; sa(c_q2,"center")
        c_l2=ws_d.cell(i,9,b["price"]); c_l2.number_format=INR; sa(c_l2,"right")
        c_d2=ws_d.cell(i,10,eff_disc(b)); c_d2.number_format=PCT; sa(c_d2,"center")
        c_nr2=ws_d.cell(i,11,f"=I{i}*(1-J{i}/100)"); c_nr2.number_format=INR; c_nr2.font=Font(color="10B981",name="Calibri",size=10,italic=True); sa(c_nr2,"right")
        c_a2=ws_d.cell(i,12,f"=H{i}*K{i}"); c_a2.number_format=INR; sa(c_a2,"right")
        ws_d.cell(i,13,b["line"])
        for c in range(1,14): ws_d.cell(i,c).border=row_border
        if i%2==0:
            for c in [1,3,4,5,6,7,8,9,10,11,12,13]: ws_d.cell(i,c).fill=alt_fill
    dl=len(boq)+1; dgt=dl+2
    ws_d.cell(dgt,4,"GRAND TOTAL").font=Font(bold=True,name="Calibri",size=11)
    c_ga2=ws_d.cell(dgt,12,f"=SUM(L2:L{dl})"); c_ga2.number_format=INR; c_ga2.font=tot_font; c_ga2.fill=tot_fill; sa(c_ga2,"right")

    # ── Per-shaft sheets ──
    sg=shaft_groups()
    for sh,items in sorted(sg.items()):
        ws2=wb.create_sheet(sh[:31]); ws2.sheet_view.showGridLines=False
        fg_hex=("1D4ED8" if sh.startswith("SH") else "065F46" if sh.startswith("K") else "475569")
        sh_fill=PatternFill("solid",fgColor=fg_hex)
        write_header(ws2,["Sr.No","Item Code","Description","Type","DN","Unit","Qty","List Price (₹)","Disc %","Net Rate (₹)","Amount (₹)"],
                     [6,22,42,14,6,5,6,14,7,14,14],fill=sh_fill)
        for i,b in enumerate(items,2):
            ws2.cell(i,1,i-1); sa(ws2.cell(i,1),"center")
            ws2.cell(i,2,b["code"]); ws2.cell(i,3,b["desc"]); ws2.cell(i,4,b["sub"])
            ws2.cell(i,5,b["dn"] if b["dn"] else "-"); sa(ws2.cell(i,5),"center")
            ws2.cell(i,6,"Nos"); sa(ws2.cell(i,6),"center")
            c_q3=ws2.cell(i,7,b["qty"]); c_q3.number_format=INT; sa(c_q3,"center")
            c_l3=ws2.cell(i,8,b["price"]); c_l3.number_format=INR; sa(c_l3,"right")
            c_d3=ws2.cell(i,9,eff_disc(b)); c_d3.number_format=PCT; sa(c_d3,"center")
            c_nr3=ws2.cell(i,10,f"=H{i}*(1-I{i}/100)"); c_nr3.number_format=INR; c_nr3.font=Font(color="10B981",name="Calibri",size=10,italic=True); sa(c_nr3,"right")
            c_a3=ws2.cell(i,11,f"=G{i}*J{i}"); c_a3.number_format=INR; sa(c_a3,"right")
            for c in range(1,12): ws2.cell(i,c).border=row_border
            if i%2==0:
                for c in range(1,12): ws2.cell(i,c).fill=alt_fill
        sl=len(items)+1; sr=sl+2
        ws2.cell(sr,3,f"SUBTOTAL — {sh}").font=Font(bold=True,name="Calibri",size=11)
        c_sq=ws2.cell(sr,7,f"=SUM(G2:G{sl})"); c_sq.number_format=INT; c_sq.font=Font(bold=True,color=fg_hex,name="Calibri",size=12); sa(c_sq,"center")
        c_sa=ws2.cell(sr,11,f"=SUM(K2:K{sl})"); c_sa.number_format=INR; c_sa.font=Font(bold=True,color=fg_hex,name="Calibri",size=13); c_sa.fill=PatternFill("solid",fgColor="F8FAFC"); sa(c_sa,"right")

    buf=BytesIO(); wb.save(buf); buf.seek(0); return buf

# ── EXCEL IMPORT ──
def import_excel(uploaded):
    wb=openpyxl.load_workbook(uploaded)
    ws=wb["Full BOQ"] if "Full BOQ" in wb.sheetnames else wb.active
    rows=list(ws.iter_rows(min_row=2,values_only=True))
    added=0
    for row in rows:
        if not row[1] or str(row[3] or "").startswith("─"): continue
        code=str(row[1]).strip()
        prod=next((p for p in PRODUCTS if p["code"]==code),None)
        shaft=str(row[11] if row[11] else row[1] or "NA").split(",")[0].strip()
        qty=int(row[6] or row[7] or 1)
        disc=float(row[9]) if row[9] is not None else None
        if prod:
            item={**prod,"qty":qty,"disc":disc,"shaft":shaft,"_id":str(uuid.uuid4())[:8]}
        else:
            item={"code":code,"desc":str(row[2] or ""),"type":"Imported","sub":str(row[3] or ""),"dn":int(row[4]) if row[4] and str(row[4]).isdigit() else 0,"price":float(row[7] or 0),"line":str(row[12] or "HT Pro"),"qty":qty,"disc":disc,"shaft":shaft,"_id":str(uuid.uuid4())[:8]}
        st.session_state.boq.append(item); added+=1
    return added

# ══════════════════════════════════════════════════════
#  UI
# ══════════════════════════════════════════════════════

# ── HEADER ──
hc1,hc2,hc3=st.columns([2,5,3])
with hc1:
    st.markdown("""<div style="background:linear-gradient(135deg,#0F172A,#1E3A5F);padding:12px 18px;border-radius:12px;color:white;">
    <span style="font-size:20px;font-weight:900;color:#FBBF24;">⚙ Huliot BOQ</span>
    </div>""",unsafe_allow_html=True)
with hc2:
    st.session_state.proj=st.text_input("proj",placeholder="Project / Site name...",label_visibility="collapsed",value=st.session_state.proj,key="proj_inp")
with hc3:
    c1,c2=st.columns(2)
    with c1:
        if st.button("🔶 HT Pro",use_container_width=True,type="primary" if st.session_state.line=="HT Pro" else "secondary"):
            st.session_state.line="HT Pro"; st.session_state.dn_f=None; st.session_state.cat_f=None; st.rerun()
    with c2:
        if st.button("🔷 Ultra Silent",use_container_width=True,type="primary" if st.session_state.line=="Ultra Silent" else "secondary"):
            st.session_state.line="Ultra Silent"; st.session_state.dn_f=None; st.session_state.cat_f=None; st.rerun()

st.markdown("<div style='height:6px'></div>",unsafe_allow_html=True)

# ── GLOBAL SHAFT BAR ──
tc,bc,_=sh_colors(st.session_state.global_sh)
sh_cnt=boq_sh_counts()
st.markdown(f"""<div style="background:{bc};border:2px solid {tc}44;border-radius:12px;padding:9px 16px;margin:2px 0 6px 0;">
<span style="font-size:13px;font-weight:800;color:{tc};">📍 ADDING TO LOCATION — All items tagged to selected location</span>
</div>""",unsafe_allow_html=True)

bar1,bar2=st.columns([2,6])
with bar1:
    new_sh=st.selectbox("loc",SH_CODES,index=SH_CODES.index(st.session_state.global_sh),format_func=sh_label,label_visibility="collapsed",key="sh_select")
    if new_sh!=st.session_state.global_sh:
        st.session_state.global_sh=new_sh; st.rerun()
with bar2:
    st.markdown("<div style='font-size:11px;color:#64748B;font-weight:600;margin-bottom:2px;'>Quick select:</div>",unsafe_allow_html=True)
    qcols=st.columns(len(QUICK_SH))
    for i,code in enumerate(QUICK_SH):
        with qcols[i]:
            cnt=sh_cnt.get(code,0)
            lbl=f"{code}\n({cnt})" if cnt else code
            if st.button(lbl,key=f"qsh_{code}",type="primary" if st.session_state.global_sh==code else "secondary",use_container_width=True):
                st.session_state.global_sh=code; st.rerun()

# ── TABS ──
filtered=get_filtered()
boq_len=len(st.session_state.boq)
gt=grand()
tab_catalog,tab_boq,tab_summary=st.tabs([
    f"📦 Catalog ({len(filtered)})",
    f"📋 BOQ ({boq_len} items)" + (f"  ·  {fmt(gt)}" if gt else ""),
    f"📊 Shaft Summary ({len(set(b.get('shaft','NA') for b in st.session_state.boq))} locations)" if boq_len else "📊 Shaft Summary"
])

# ══ CATALOG ══
with tab_catalog:
    st.markdown("<div style='font-size:10px;font-weight:700;color:#6B7280;text-transform:uppercase;margin:4px 0 6px 0;'>Select Size (DN)</div>",unsafe_allow_html=True)
    dn_counts=get_dn_counts()
    dn_cols=st.columns(len(DN_LIST)+1)
    with dn_cols[0]:
        if st.button(f"ALL\n{len([p for p in PRODUCTS if p['line']==st.session_state.line])}",use_container_width=True,type="primary" if st.session_state.dn_f is None else "secondary"):
            st.session_state.dn_f=None; st.rerun()
    for i,d in enumerate(DN_LIST):
        cnt=dn_counts.get(d,0)
        if not cnt: continue
        with dn_cols[i+1]:
            if st.button(f"DN{d}\n{cnt}",use_container_width=True,type="primary" if st.session_state.dn_f==d else "secondary",key=f"dn_{d}"):
                st.session_state.dn_f=None if st.session_state.dn_f==d else d; st.rerun()

    cat_cols=st.columns(len(CATS)+1)
    with cat_cols[0]:
        if st.button("All Types",use_container_width=True,type="primary" if not st.session_state.cat_f else "secondary"):
            st.session_state.cat_f=None; st.rerun()
    for i,c in enumerate(CATS):
        with cat_cols[i+1]:
            if st.button(f"{CAT_ICONS.get(c,'')} {c}",use_container_width=True,type="primary" if st.session_state.cat_f==c else "secondary",key=f"cat_{c}"):
                st.session_state.cat_f=None if st.session_state.cat_f==c else c; st.rerun()

    search_val=st.text_input("search",placeholder="🔍  Search item code or description...",label_visibility="collapsed",value=st.session_state.srch,key="srch_inp")
    if search_val!=st.session_state.srch:
        st.session_state.srch=search_val; st.rerun()

    fi1,fi2=st.columns([5,1])
    with fi1:
        info=[]
        if st.session_state.dn_f: info.append(f"DN{st.session_state.dn_f}")
        if st.session_state.cat_f: info.append(st.session_state.cat_f)
        if st.session_state.srch: info.append(f'"{st.session_state.srch}"')
        st.markdown(f"<div style='font-weight:700;font-size:13px;padding:4px 0;'>{len(filtered)} items{('  ·  Filter: '+', '.join(info)) if info else ''}</div>",unsafe_allow_html=True)
    with fi2:
        if info:
            if st.button("✕ Clear",use_container_width=True):
                st.session_state.dn_f=None; st.session_state.cat_f=None; st.session_state.srch=""; st.rerun()

    # Table header
    h1,h2,h3,h4,h5,h6,h7=st.columns([3,5,1,2,2,1,1.5])
    for hdr,col in zip(["CODE","DESCRIPTION","DN","TYPE","LIST PRICE","QTY","ADD"],[h1,h2,h3,h4,h5,h6,h7]):
        col.markdown(f"<div style='font-size:10px;font-weight:700;color:#64748B;text-transform:uppercase;background:#F8FAFC;padding:6px 4px;border-bottom:2px solid #E2E8F0;'>{hdr}</div>",unsafe_allow_html=True)

    if not filtered:
        st.info("No products found. Try clearing filters.")
    else:
        for prod in filtered:
            c1,c2,c3,c4,c5,c6,c7=st.columns([3,5,1,2,2,1,1.5])
            with c1:
                st.markdown(f"<div style='font-family:monospace;font-size:10px;color:#64748B;padding:4px 2px;'>{prod['code']}</div>",unsafe_allow_html=True)
            with c2:
                st.markdown(f"<div style='font-weight:600;font-size:12px;color:#1E293B;padding:2px;'>{prod['desc']}<br><span style='font-size:10px;color:#94A3B8;'>{prod['sub']}</span></div>",unsafe_allow_html=True)
            with c3:
                if prod["dn"]>0:
                    dc=DN_COLORS.get(prod["dn"],"#64748B")
                    st.markdown(f"<div style='text-align:center;'><span style='background:{dc}18;color:{dc};padding:2px 7px;border-radius:10px;font-size:11px;font-weight:800;'>{prod['dn']}</span></div>",unsafe_allow_html=True)
            with c4:
                _cc={"Pipe":"#1D4ED8","Bend":"#B45309","Branch":"#065F46","Trap":"#6D28D9","Coupler":"#0E7490","Reducer":"#3730A3","Inspection":"#92400E","Clamp":"#991B1B","Accessory":"#374151"}
                _cb={"Pipe":"#DBEAFE","Bend":"#FEF3C7","Branch":"#D1FAE5","Trap":"#EDE9FE","Coupler":"#CFFAFE","Reducer":"#E0E7FF","Inspection":"#FEF3C7","Clamp":"#FEE2E2","Accessory":"#F3F4F6"}
                cc=_cc.get(prod["type"],"#64748B"); ccb=_cb.get(prod["type"],"#F1F5F9")
                st.markdown(f"<div style='text-align:center;'><span style='background:{ccb};color:{cc};padding:2px 7px;border-radius:10px;font-size:10px;font-weight:700;'>{prod['type']}</span></div>",unsafe_allow_html=True)
            with c5:
                st.markdown(f"<div style='text-align:right;font-weight:700;font-size:13px;color:#1E293B;padding:4px 4px;'>{fmt(prod['price'])}</div>",unsafe_allow_html=True)
            with c6:
                qty=st.number_input("",min_value=1,value=1,step=1,label_visibility="collapsed",key=f"qty_{prod['code']}")
            with c7:
                tc2,_,_=sh_colors(st.session_state.global_sh)
                if st.button(f"+ Add\n{st.session_state.global_sh}",key=f"add_{prod['code']}",use_container_width=True,type="primary"):
                    add_to_boq(prod,qty)
                    st.toast(f"✅ {prod['desc']} → {st.session_state.global_sh}",icon="✅")
                    st.rerun()
            st.markdown("<hr style='margin:0;border:none;border-top:1px solid #F1F5F9;'>",unsafe_allow_html=True)

    if boq_len:
        st.markdown("<div style='height:8px'></div>",unsafe_allow_html=True)
        if st.button(f"📋 View BOQ — {boq_len} items · {fmt(gt)}",type="primary",use_container_width=True):
            st.rerun()

# ══ BOQ ══
with tab_boq:
    if not st.session_state.boq:
        st.info("BOQ is empty. Go to Catalog and add items.")
    else:
        # Controls
        bc1,bc2,bc3,bc4,bc5=st.columns([2,1.5,1.5,1.5,1])
        with bc1:
            st.markdown(f"<div style='font-weight:700;font-size:15px;padding-top:8px;'>{st.session_state.proj or 'Huliot BOQ'}</div>",unsafe_allow_html=True)
        with bc2:
            g_d=st.number_input("Global Discount %",min_value=0.0,max_value=100.0,value=float(st.session_state.g_disc),step=0.5,key="g_disc_inp")
            st.session_state.g_disc=g_d
        with bc3:
            xl=export_excel()
            pn=(st.session_state.proj or "Huliot").replace(" ","_")
            st.download_button("⬇ Export Excel",data=xl,file_name=f"{pn}_BOQ.xlsx",mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",use_container_width=True,type="primary")
        with bc4:
            uploaded=st.file_uploader("Import",type=["xlsx","xls"],label_visibility="collapsed",key="xl_up")
            if uploaded:
                n=import_excel(uploaded); st.toast(f"✅ Imported {n} items",icon="✅"); st.rerun()
        with bc5:
            if st.button("🗑 Clear",use_container_width=True):
                st.session_state.boq=[]; st.rerun()

        st.markdown("<div style='height:4px'></div>",unsafe_allow_html=True)

        # ── Copy Shaft ──
        with st.expander("📋 Copy Shaft Quantities"):
            cp1,cp2,cp3,cp4=st.columns([2.5,2.5,1.5,2])
            used_shafts=sorted(set(b.get("shaft","NA") for b in st.session_state.boq),key=lambda x:("" if x=="NA" else x))
            with cp1:
                st.markdown("<div style='font-size:11px;font-weight:700;color:#1D4ED8;margin-bottom:4px;'>📤 Copy FROM (existing shaft)</div>",unsafe_allow_html=True)
                copy_from=st.selectbox("from",["— Select —"]+used_shafts,label_visibility="collapsed",key="cp_from",format_func=lambda x:x if x=="— Select —" else sh_label(x))
            with cp2:
                st.markdown("<div style='font-size:11px;font-weight:700;color:#065F46;margin-bottom:4px;'>📥 Copy TO (destination)</div>",unsafe_allow_html=True)
                copy_to=st.selectbox("to",SH_CODES,label_visibility="collapsed",key="cp_to",format_func=sh_label)
            with cp3:
                st.markdown("<div style='margin-bottom:4px;height:19px'></div>",unsafe_allow_html=True)
                do_copy=st.button("⧉ Copy",use_container_width=True,type="primary",key="cp_btn")
            with cp4:
                sg=shaft_groups()
                if copy_from and copy_from!="— Select —":
                    from_items=sg.get(copy_from,[])
                    src_total=sum(amt(b) for b in from_items)
                    st.markdown(f"""<div style='background:#F0FDF4;border:1px solid #A7F3D0;border-radius:8px;padding:8px 12px;font-size:12px;'>
                    <b>{len(from_items)} items</b> · {fmt(src_total)}<br>
                    <span style='color:#64748B;'>{copy_from} → {copy_to}</span></div>""",unsafe_allow_html=True)
            if do_copy:
                if not copy_from or copy_from=="— Select —":
                    st.warning("⚠ Select a source shaft first.")
                elif copy_from==copy_to:
                    st.warning("⚠ Source and destination are the same.")
                else:
                    from_items=sg.get(copy_from,[])
                    if not from_items:
                        st.warning(f"⚠ No items in {copy_from}.")
                    else:
                        added,merged=0,0
                        for item in from_items:
                            ex=next((b for b in st.session_state.boq if b["code"]==item["code"] and b.get("shaft")==copy_to),None)
                            if ex: ex["qty"]+=item["qty"]; merged+=1
                            else: st.session_state.boq.append({**item,"shaft":copy_to,"_id":str(uuid.uuid4())[:8]}); added+=1
                        st.success(f"✅ Copied to {copy_to} — {added} new items, {merged} merged")
                        st.rerun()

        st.markdown("<div style='height:4px'></div>",unsafe_allow_html=True)

        # ── Shaft filter chips ──
        fsh_cols=st.columns(min(len(used_shafts)+1,10))
        with fsh_cols[0]:
            if st.button(f"All ({boq_len})",use_container_width=True,type="primary" if st.session_state.boq_fsh=="ALL" else "secondary",key="fsh_all"):
                st.session_state.boq_fsh="ALL"; st.rerun()
        for i,sh in enumerate(used_shafts[:9]):
            with fsh_cols[i+1]:
                cnt=sh_cnt.get(sh,0)
                if st.button(f"{sh} ({cnt})",use_container_width=True,type="primary" if st.session_state.boq_fsh==sh else "secondary",key=f"fsh_{sh}"):
                    st.session_state.boq_fsh=sh; st.rerun()

        is_all=st.session_state.boq_fsh=="ALL"
        if is_all:
            display_rows=agg_boq()
        else:
            display_rows=[b for b in st.session_state.boq if b.get("shaft","NA")==st.session_state.boq_fsh]
        display_total=sum(r["qty"]*net_rate(r) for r in display_rows)

        lbl=f"GRAND TOTAL — All Shafts Combined" if is_all else f"TOTAL — {st.session_state.boq_fsh}"
        st.markdown(f"<div style='text-align:right;font-weight:800;color:#1D4ED8;font-size:14px;padding:4px 0;'>{lbl}: {fmt(display_total)}</div>",unsafe_allow_html=True)

        # Table header
        if is_all:
            bh=st.columns([0.4,2.5,4,0.8,1.2,1.5,0.8,1.5,1.8,2.2])
            for hdr,col in zip(["#","CODE","DESCRIPTION","DN","TOTAL QTY","LIST PRICE","DISC%","NET RATE","AMOUNT","SHAFTS"],bh):
                col.markdown(f"<div style='font-size:9px;font-weight:700;color:#64748B;text-transform:uppercase;background:#F8FAFC;padding:5px 3px;border-bottom:2px solid #E2E8F0;text-align:center;'>{hdr}</div>",unsafe_allow_html=True)
        else:
            bh=st.columns([0.4,1.2,2.5,4,0.8,1,1.5,0.8,1.5,1.8,0.5])
            for hdr,col in zip(["#","LOC","CODE","DESCRIPTION","DN","QTY","LIST PRICE","DISC%","NET RATE","AMOUNT","DEL"],bh):
                col.markdown(f"<div style='font-size:9px;font-weight:700;color:#64748B;text-transform:uppercase;background:#F8FAFC;padding:5px 3px;border-bottom:2px solid #E2E8F0;text-align:center;'>{hdr}</div>",unsafe_allow_html=True)

        for idx,row in enumerate(display_rows):
            nr=net_rate(row); a=row["qty"]*nr
            dc=DN_COLORS.get(row["dn"],"#64748B")
            if is_all:
                r=st.columns([0.4,2.5,4,0.8,1.2,1.5,0.8,1.5,1.8,2.2])
                with r[0]: st.markdown(f"<div style='text-align:center;color:#94A3B8;font-size:11px;padding:8px 0;'>{idx+1}</div>",unsafe_allow_html=True)
                with r[1]: st.markdown(f"<div style='font-family:monospace;font-size:9px;color:#64748B;padding:6px 2px;'>{row['code']}</div>",unsafe_allow_html=True)
                with r[2]: st.markdown(f"<div style='font-weight:600;font-size:12px;color:#1E293B;padding:3px 2px;'>{row['desc']}<br><span style='font-size:10px;color:#94A3B8;'>{row['sub']}</span></div>",unsafe_allow_html=True)
                with r[3]:
                    if row["dn"]>0: st.markdown(f"<div style='text-align:center;'><span style='background:{dc}18;color:{dc};padding:2px 6px;border-radius:8px;font-size:11px;font-weight:800;'>{row['dn']}</span></div>",unsafe_allow_html=True)
                with r[4]: st.markdown(f"<div style='text-align:center;'><span style='background:#DBEAFE;color:#1D4ED8;padding:4px 10px;border-radius:8px;font-size:14px;font-weight:900;'>{row['qty']}</span></div>",unsafe_allow_html=True)
                with r[5]: st.markdown(f"<div style='text-align:right;font-size:12px;color:#64748B;padding:6px 4px;'>{fmt(row['price'])}</div>",unsafe_allow_html=True)
                with r[6]: st.markdown(f"<div style='text-align:center;font-size:12px;color:#64748B;padding:6px 4px;'>{eff_disc(row)}%</div>",unsafe_allow_html=True)
                with r[7]: st.markdown(f"<div style='text-align:right;font-weight:700;color:#10B981;font-size:12px;padding:6px 4px;'>{fmt(nr)}</div>",unsafe_allow_html=True)
                with r[8]: st.markdown(f"<div style='text-align:right;font-weight:800;font-size:13px;color:#1E293B;padding:6px 4px;'>{fmt(a)}</div>",unsafe_allow_html=True)
                with r[9]:
                    shafts=row.get("shafts",[row.get("shaft","NA")])
                    badges="".join(f"<span style='background:{sh_colors(s)[1]};color:{sh_colors(s)[0]};border:1px solid {sh_colors(s)[0]}44;padding:1px 6px;border-radius:6px;font-size:10px;font-weight:700;margin:1px;display:inline-block;'>{s}</span>" for s in sorted(shafts))
                    st.markdown(f"<div style='padding:4px 2px;line-height:1.8;'>{badges}</div>",unsafe_allow_html=True)
            else:
                b=row
                r=st.columns([0.4,1.2,2.5,4,0.8,1,1.5,0.8,1.5,1.8,0.5])
                with r[0]: st.markdown(f"<div style='text-align:center;color:#94A3B8;font-size:11px;padding:8px 0;'>{idx+1}</div>",unsafe_allow_html=True)
                with r[1]:
                    new_sh=st.selectbox("sh",SH_CODES,index=SH_CODES.index(b.get("shaft","NA")),label_visibility="collapsed",key=f"bsh_{b['_id']}",format_func=lambda x:x)
                    if new_sh!=b.get("shaft","NA"): b["shaft"]=new_sh; st.rerun()
                with r[2]: st.markdown(f"<div style='font-family:monospace;font-size:9px;color:#64748B;padding:6px 2px;'>{b['code']}</div>",unsafe_allow_html=True)
                with r[3]: st.markdown(f"<div style='font-weight:600;font-size:12px;color:#1E293B;padding:3px 2px;'>{b['desc']}<br><span style='font-size:10px;color:#94A3B8;'>{b['sub']}</span></div>",unsafe_allow_html=True)
                with r[4]:
                    if b["dn"]>0: st.markdown(f"<div style='text-align:center;'><span style='background:{dc}18;color:{dc};padding:2px 6px;border-radius:8px;font-size:11px;font-weight:800;'>{b['dn']}</span></div>",unsafe_allow_html=True)
                with r[5]:
                    nq=st.number_input("qty",min_value=1,value=b["qty"],step=1,label_visibility="collapsed",key=f"bqty_{b['_id']}")
                    if nq!=b["qty"]: b["qty"]=nq; st.rerun()
                with r[6]: st.markdown(f"<div style='text-align:right;font-size:12px;color:#64748B;padding:6px 4px;'>{fmt(b['price'])}</div>",unsafe_allow_html=True)
                with r[7]:
                    nd=st.number_input("disc",min_value=0.0,max_value=100.0,value=float(b["disc"]) if b["disc"] is not None else float(st.session_state.g_disc),step=0.5,label_visibility="collapsed",key=f"bdisc_{b['_id']}")
                    if nd!=(b["disc"] if b["disc"] is not None else st.session_state.g_disc): b["disc"]=nd; st.rerun()
                with r[8]: st.markdown(f"<div style='text-align:right;font-weight:700;color:#10B981;font-size:12px;padding:6px 4px;'>{fmt(net_rate(b))}</div>",unsafe_allow_html=True)
                with r[9]: st.markdown(f"<div style='text-align:right;font-weight:800;font-size:13px;color:#1E293B;padding:6px 4px;'>{fmt(amt(b))}</div>",unsafe_allow_html=True)
                with r[10]:
                    if st.button("✕",key=f"del_{b['_id']}",use_container_width=True):
                        st.session_state.boq=[x for x in st.session_state.boq if x["_id"]!=b["_id"]]; st.rerun()
            st.markdown("<hr style='margin:0;border:none;border-top:1px solid #F1F5F9;'>",unsafe_allow_html=True)

        st.markdown(f"""<div style="background:linear-gradient(90deg,#0F172A,#1E3A5F);color:white;padding:10px 14px;border-radius:0 0 10px 10px;display:flex;justify-content:space-between;align-items:center;">
        <div><span style="font-weight:700;font-size:14px;">{lbl}</span><br>
        <span style="font-size:11px;opacity:.6;">{""+str(boq_len)+" entries · "+str(len(used_shafts))+" locations" if is_all else str(len(display_rows))+" items"}</span></div>
        <span style="font-weight:900;font-size:20px;color:#FBBF24;">{fmt(display_total)}</span></div>""",unsafe_allow_html=True)
        st.markdown("<div style='padding:5px 12px;background:#FFFBEB;border-top:1px solid #FDE68A;font-size:10px;color:#92400E;'>⚠ Prices ex-factory/depot · GST extra · W.E.F April 2026 · Excel has live formulas</div>",unsafe_allow_html=True)

# ══ SUMMARY ══
with tab_summary:
    sg=shaft_groups()
    sh_sorted=sorted(sg.keys(),key=lambda x:("" if x=="NA" else x))
    if not sh_sorted:
        st.info("No data yet. Add items to BOQ first.")
    else:
        total_sh=sum(sum(amt(b) for b in sg[sh]) for sh in sh_sorted if sh.startswith("SH"))
        total_k=sum(sum(amt(b) for b in sg[sh]) for sh in sh_sorted if sh.startswith("K"))
        total_na=sum(amt(b) for b in sg.get("NA",[]))
        sc1,sc2,sc3,sc4=st.columns(4)
        sc1.markdown(f"""<div style="background:linear-gradient(135deg,#0F172A,#1E3A5F);border-radius:12px;padding:16px;color:white;">
        <div style="font-size:11px;opacity:.7;margin-bottom:4px;">GRAND TOTAL</div>
        <div style="font-size:26px;font-weight:900;color:#FBBF24;">{fmt(gt)}</div>
        <div style="font-size:11px;opacity:.6;">{boq_len} items · {len(sh_sorted)} locations</div></div>""",unsafe_allow_html=True)
        sc2.markdown(f"""<div style="background:#EFF6FF;border:2px solid #BFDBFE;border-radius:12px;padding:16px;">
        <div style="font-size:11px;color:#1D4ED8;font-weight:700;margin-bottom:4px;">SHAFT TOTAL</div>
        <div style="font-size:22px;font-weight:900;color:#1D4ED8;">{fmt(total_sh)}</div>
        <div style="font-size:11px;color:#3B82F6;">{len([s for s in sh_sorted if s.startswith('SH')])} shafts</div></div>""",unsafe_allow_html=True)
        sc3.markdown(f"""<div style="background:#F0FDF4;border:2px solid #A7F3D0;border-radius:12px;padding:16px;">
        <div style="font-size:11px;color:#065F46;font-weight:700;margin-bottom:4px;">KITCHEN TOTAL</div>
        <div style="font-size:22px;font-weight:900;color:#065F46;">{fmt(total_k)}</div>
        <div style="font-size:11px;color:#10B981;">{len([s for s in sh_sorted if s.startswith('K')])} kitchens</div></div>""",unsafe_allow_html=True)
        sc4.markdown(f"""<div style="background:#F8FAFC;border:2px solid #CBD5E1;border-radius:12px;padding:16px;">
        <div style="font-size:11px;color:#64748B;font-weight:700;margin-bottom:4px;">UNASSIGNED</div>
        <div style="font-size:22px;font-weight:900;color:#64748B;">{fmt(total_na)}</div>
        <div style="font-size:11px;color:#94A3B8;">{len(sg.get('NA',[]))} items</div></div>""",unsafe_allow_html=True)

        st.markdown("<div style='height:10px'></div>",unsafe_allow_html=True)

        for sh in sh_sorted:
            items=sg[sh]; sub=sum(amt(b) for b in items)
            pct=round(sub/gt*100,1) if gt else 0
            tc4,bc4,brd4=sh_colors(sh)
            sh_name=("Unassigned" if sh=="NA" else f"Shaft {int(sh[2:])}" if sh.startswith("SH") else f"Kitchen / Flat {int(sh[1:])}")
            st.markdown(f"""<div style="background:{bc4};border-left:5px solid {brd4};padding:10px 14px;border-radius:10px 10px 0 0;display:flex;justify-content:space-between;align-items:center;margin-top:8px;">
            <div style="display:flex;align-items:center;gap:10px;">
            <span style="background:{tc4};color:white;padding:4px 12px;border-radius:8px;font-size:13px;font-weight:900;">{sh}</span>
            <span style="font-weight:700;color:#1E293B;font-size:14px;">{sh_name}</span>
            <span style="font-size:11px;color:#64748B;">{len(items)} items</span>
            <span style="font-size:11px;background:rgba(0,0,0,.08);padding:2px 8px;border-radius:10px;color:#475569;">{pct}%</span>
            </div><div style="font-size:18px;font-weight:900;color:{tc4};">{fmt(sub)}</div></div>""",unsafe_allow_html=True)
            st.markdown(f"""<div style="height:5px;background:#F1F5F9;">
            <div style="height:100%;width:{pct}%;background:{tc4};"></div></div>""",unsafe_allow_html=True)

            tbl="<table style='width:100%;border-collapse:collapse;font-size:11px;'>"
            tbl+="<thead><tr style='background:#F8FAFC;'>"
            for h in ["#","Code","Description","DN","Qty","Net Rate","Amount"]:
                align="right" if h in ["Net Rate","Amount"] else "center" if h in ["#","DN","Qty"] else "left"
                tbl+=f"<th style='padding:5px 8px;font-size:9px;font-weight:700;color:#64748B;text-transform:uppercase;border-bottom:1px solid #E2E8F0;text-align:{align};'>{h}</th>"
            tbl+="</tr></thead><tbody>"
            for i,b in enumerate(items):
                bg="white" if i%2==0 else "#FAFBFD"
                dc2=DN_COLORS.get(b["dn"],"#64748B")
                dn_b=f"<span style='background:{dc2}18;color:{dc2};padding:1px 6px;border-radius:8px;font-size:11px;font-weight:800;'>{b['dn']}</span>" if b["dn"] else ""
                tbl+=f"<tr style='background:{bg};border-bottom:1px solid #F8FAFC;'>"
                tbl+=f"<td style='padding:5px 8px;text-align:center;color:#94A3B8;'>{i+1}</td>"
                tbl+=f"<td style='padding:5px 8px;font-family:monospace;font-size:9px;color:#64748B;'>{b['code']}</td>"
                tbl+=f"<td style='padding:5px 8px;'><span style='font-weight:600;color:#1E293B;'>{b['desc']}</span><br><span style='font-size:10px;color:#94A3B8;'>{b['sub']}</span></td>"
                tbl+=f"<td style='padding:5px 8px;text-align:center;'>{dn_b}</td>"
                tbl+=f"<td style='padding:5px 8px;text-align:center;font-weight:700;'>×{b['qty']}</td>"
                tbl+=f"<td style='padding:5px 8px;text-align:right;color:#10B981;font-weight:600;'>{fmt(net_rate(b))}</td>"
                tbl+=f"<td style='padding:5px 8px;text-align:right;font-weight:800;color:#1E293B;'>{fmt(amt(b))}</td></tr>"
            tbl+=f"</tbody><tfoot><tr style='background:{bc4};'><td colspan='5' style='padding:7px 10px;font-weight:700;color:{tc4};font-size:12px;'>Subtotal — {sh}</td>"
            tbl+=f"<td colspan='2' style='padding:7px 10px;text-align:right;font-weight:900;font-size:14px;color:{tc4};'>{fmt(sub)}</td></tr></tfoot></table>"
            st.markdown(tbl,unsafe_allow_html=True)

        st.markdown("<div style='height:8px'></div>",unsafe_allow_html=True)
        xl2=export_excel()
        pn2=(st.session_state.proj or "Huliot").replace(" ","_")
        st.download_button("⬇ Export Full Excel (Shaft-wise Sheets + Formulas)",data=xl2,file_name=f"{pn2}_BOQ_ShaftWise.xlsx",mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",use_container_width=True,type="primary")
